"""
masterfile_hreflang.py
Hreflang Masterfile - 8 sheets, xlsxwriter + openpyxl two-pass.
Adapted from working Jupyter reference. Uses app.services.rulebook for URL classification.
"""
import os, io, math, tempfile
import pandas as pd
import xlsxwriter
import xlsxwriter.utility
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.rulebook import load_rulebook, classify_url as _rulebook_classify

RED   = "#FF0000"
WHITE = "#FFFFFF"
BLACK = "#000000"
GRAY  = "#D9D9D9"
PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2, "N/A": 3}


def _classify(url, rulebook):
    if not url or url == "-":
        return ("-", "-", "-", "N/A")
    try:
        t1, t2, lang, pri = _rulebook_classify(url, rulebook)
        return (t1 or "-", t2 or "-", lang or "-", pri or "N/A")
    except Exception:
        return ("-", "-", "-", "N/A")


def safe_num(v):
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except Exception:
        return None


def _clean(v):
    s = str(v).strip() if v is not None else ""
    return "-" if (not s or s in ("nan", "None", "NaN")) else s


def _gc(df, *names):
    for n in names:
        m = next((c for c in df.columns if c.lower() == n.lower()), None)
        if m:
            return m
    return None


def load_csv(folder, filename):
    p = os.path.join(folder, filename)
    if os.path.exists(p) and os.path.getsize(p) > 10:
        try:
            return pd.read_csv(p, encoding="utf-8", low_memory=False)
        except Exception:
            try:
                return pd.read_csv(p, encoding="latin-1", low_memory=False)
            except Exception:
                pass
    return pd.DataFrame()


def filter_indexable_200_html(df):
    if df.empty:
        return df
    out = df.copy()
    idx = _gc(out, "Indexability")
    sc  = _gc(out, "Status Code")
    ct  = _gc(out, "Content Type")
    if idx:
        out = out[out[idx].astype(str).str.strip().str.lower() == "indexable"]
    if sc:
        out = out[out[sc].astype(str).str.strip() == "200"]
    if ct:
        out = out[out[ct].astype(str).str.contains("text/html", case=False, na=False)]
    return out


def build_internal_map(df_int):
    m = {}
    if df_int.empty:
        return m
    a    = _gc(df_int, "Address")
    sc   = _gc(df_int, "Status Code")
    idx  = _gc(df_int, "Indexability")
    idxs = _gc(df_int, "Indexability Status")
    if not a:
        return m
    for _, r in df_int.iterrows():
        url = _clean(r[a])
        m[url] = {
            "status_code":         _clean(r.get(sc,   "")) if sc   else "-",
            "indexability":        _clean(r.get(idx,  "")) if idx  else "-",
            "indexability_status": _clean(r.get(idxs, "")) if idxs else "-",
        }
    return m


def build_gsc_map(df_gsc):
    m = {}
    if df_gsc.empty:
        return m
    a   = _gc(df_gsc, "Address")
    imp = next((c for c in df_gsc.columns if "impression" in c.lower()), None)
    clk = next((c for c in df_gsc.columns if "click" in c.lower()), None)
    if a:
        for _, r in df_gsc.iterrows():
            m[_clean(r[a])] = {
                "impressions": safe_num(r.get(imp)) if imp else None,
                "clicks":      safe_num(r.get(clk)) if clk else None,
            }
    return m


def build_ga_map(df_ga):
    m = {}
    if df_ga.empty:
        return m
    a = _gc(df_ga, "Address")
    s = next((c for c in df_ga.columns if "session" in c.lower()), None)
    if a and s:
        for _, r in df_ga.iterrows():
            m[_clean(r[a])] = safe_num(r.get(s))
    return m


def sorted_themes_from_rows(rows, issue_keys):
    theme_data = {}
    for r in rows:
        th  = r.get("Page Theme 1", "-") or "-"
        ik  = r.get("_issue_key", "")
        pri = r.get("_priority", "N/A")
        if th not in theme_data:
            theme_data[th] = {k: 0 for k in issue_keys}
            theme_data[th]["_priority"] = "N/A"
        if ik in issue_keys:
            theme_data[th][ik] += 1
        if PRIORITY_ORDER.get(pri, 3) < PRIORITY_ORDER.get(theme_data[th]["_priority"], 3):
            theme_data[th]["_priority"] = pri
    return sorted(theme_data.items(), key=lambda x: PRIORITY_ORDER.get(x[1]["_priority"], 3))


def _thin():
    return Side(style="thin")


def _border():
    return Border(left=_thin(), right=_thin(), top=_thin(), bottom=_thin())


def _gray_fill():
    return PatternFill(fill_type="solid", fgColor="FFD9D9D9")


def _white_fill():
    return PatternFill(fill_type="solid", fgColor="FFFFFFFF")


def make_formats(wb):
    def f(**kw): return wb.add_format(kw)
    return {
        "red_lft":      f(bold=True,  font_name="Calibri", font_size=8,  font_color=WHITE, bg_color=RED,   border=1, align="left",   valign="vcenter", text_wrap=True),
        "cell":         f(             font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter"),
        "cell_lft":     f(             font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="left",   valign="vcenter"),
        "t1_issue_hdr": f(bold=True,  font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter", text_wrap=True),
        "t3_hdr":       f(bold=True,  font_name="Calibri", font_size=9,  font_color=BLACK, bg_color=GRAY,  border=1, align="center", valign="vcenter", text_wrap=True),
        "t3_hdr_lft":   f(bold=True,  font_name="Calibri", font_size=9,  font_color=BLACK, bg_color=GRAY,  border=1, align="left",   valign="vcenter", text_wrap=True),
        "t3_cell":      f(             font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter"),
        "t3_cell_lft":  f(             font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="left",   valign="vcenter"),
        "num":          f(             font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter", num_format="#,##0"),
        "pct":          f(             font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter", num_format="0.00%"),
        "lbl":          f(bold=True,  font_name="Calibri", font_size=11, font_color=BLACK),
        "t2_title":     f(bold=True,  font_name="Calibri", font_size=8,  font_color=WHITE, bg_color=RED,   border=1, align="center", valign="vcenter"),
        "t2_hdr_lft":   f(bold=True,  font_name="Calibri", font_size=8,  font_color=BLACK, bg_color=GRAY,  border=1, align="left",   valign="vcenter", text_wrap=True),
        "t2_hdr_ctr":   f(bold=True,  font_name="Calibri", font_size=8,  font_color=BLACK, bg_color=GRAY,  border=1, align="center", valign="vcenter", text_wrap=True),
        "t2_hdr_wht":   f(bold=True,  font_name="Calibri", font_size=8,  font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter", text_wrap=True),
        "t2_cell":      f(             font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter"),
        "t2_cell_lft":  f(             font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="left",   valign="vcenter"),
        "summary":      f(             font_name="Calibri", font_size=10, font_color=BLACK, text_wrap=True, valign="top", border=1),
    }


# ── SHEET 1 ───────────────────────────────────────────────────────────────────
def build_sheet1(wb, fmts, folder, rulebook, internal_map, gsc_map, ga_map, total_indexable):
    ws = wb.add_worksheet("Href Lang - Combined_Group_1")
    for col, w in zip("ABCDEFGHIJKLMNO",
                      [26.27, 20.45, 15.18, 19.18, 19.45, 24.82,
                       25.0, 20.54, 22.82, 24.73, 26.45, 19.45,
                       17.82, 13.27, 16.45]):
        ws.set_column("%s:%s" % (col, col), w)

    ws.merge_range(1, 0, 10, 5, "", fmts["summary"])

    G1_ISSUES = [
        "hreflang_outside_head",
        "hreflang_missing",
        "hreflang_missing_xdefault",
        "hreflang_incorrect_language_region_codes",
    ]

    rows = []
    for label, csv_file in [
        ("hreflang_outside_head",                   "hreflang_outside_head.csv"),
        ("hreflang_missing",                         "hreflang_missing.csv"),
        ("hreflang_missing_xdefault",                "hreflang_missing_xdefault.csv"),
        ("hreflang_incorrect_language_region_codes", "hreflang_incorrect_language_region_codes.csv"),
    ]:
        df = load_csv(folder, csv_file)
        if df.empty:
            continue
        addr = _gc(df, "Address")
        h1c  = _gc(df, "HTML hreflang 1")
        h1u  = _gc(df, "HTML hreflang 1 URL")
        ht1  = _gc(df, "HTTP hreflang 1")
        ht1u = _gc(df, "HTTP hreflang 1 URL")
        s1   = _gc(df, "Sitemap hreflang 1")
        s1u  = _gc(df, "Sitemap hreflang 1 URL")
        if not addr:
            continue
        for _, row in df.iterrows():
            url = _clean(row.get(addr, ""))
            if not url or url == "-":
                continue
            int_d = internal_map.get(url, {})
            t1, t2, _, pri = _classify(url, rulebook)
            gsc = gsc_map.get(url, {})
            rows.append({
                "_issue_key": label,
                "Error Type": label,
                "Address": url,
                "Page Theme 1": t1,
                "Page Theme 2": t2,
                "Indexability":        int_d.get("indexability",        "-"),
                "Indexability Status": int_d.get("indexability_status", "-"),
                "HTML hreflang 1":        _clean(row.get(h1c,  "")) if h1c  else "-",
                "HTML hreflang 1 URL":    _clean(row.get(h1u,  "")) if h1u  else "-",
                "HTTP hreflang 1":        _clean(row.get(ht1,  "")) if ht1  else "-",
                "HTTP hreflang 1 URL":    _clean(row.get(ht1u, "")) if ht1u else "-",
                "Sitemap hreflang 1":     _clean(row.get(s1,   "")) if s1   else "-",
                "Sitemap hreflang 1 URL": _clean(row.get(s1u,  "")) if s1u  else "-",
                "Impressions":      gsc.get("impressions"),
                "Clicks":           gsc.get("clicks"),
                "Organic Sessions": ga_map.get(url),
                "_priority": pri,
            })

    for ik in G1_ISSUES:
        if not any(r["_issue_key"] == ik for r in rows):
            rows.append({
                "_issue_key": ik, "Error Type": ik, "Address": "-",
                "Page Theme 1": "-", "Page Theme 2": "-",
                "Indexability": "-", "Indexability Status": "-",
                "HTML hreflang 1": "-", "HTML hreflang 1 URL": "-",
                "HTTP hreflang 1": "-", "HTTP hreflang 1 URL": "-",
                "Sitemap hreflang 1": "-", "Sitemap hreflang 1 URL": "-",
                "Impressions": None, "Clicks": None, "Organic Sessions": None,
                "_priority": "N/A",
            })

    rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))
    themes = sorted_themes_from_rows(rows, G1_ISSUES)

    T3_HDR  = 32
    T3_DATA = 33
    T3_XL   = T3_DATA + 1

    ws.write(15, 0, "Table 1", fmts["lbl"])
    ws.set_row(16, 29)
    ws.write(16, 0, "URL Issue Types", fmts["red_lft"])
    for i, ik in enumerate(G1_ISSUES):
        ws.write(16, i + 1, ik, fmts["t1_issue_hdr"])
    ws.write(17, 0, "Issue Priority", fmts["red_lft"])
    for i in range(4):
        ws.write(17, i + 1, "High", fmts["cell"])
    ws.write(18, 0, "#Affected URLs", fmts["red_lft"])
    for i, ik in enumerate(G1_ISSUES):
        ws.write_formula(18, i + 1,
            '=COUNTIF(A%d:A1048576,"%s")' % (T3_XL, ik), fmts["num"])
    ws.set_row(19, 21)
    ws.write(19, 0, "% Share against Total  HTML URLs Crawled", fmts["red_lft"])
    for i in range(4):
        cl = xlsxwriter.utility.xl_col_to_name(i + 1)
        ws.write_formula(19, i + 1, "=%s19/%d" % (cl, total_indexable), fmts["pct"])

    ws.write(21, 0, "Table 2", fmts["lbl"])
    ws.merge_range(22, 0, 22, 5, "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(23, 29)
    ws.write(23, 0, "Page Theme 1",               fmts["t2_hdr_lft"])
    ws.write(23, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_ctr"])
    for i, ik in enumerate(G1_ISSUES):
        ws.write(23, i + 2, ik, fmts["t2_hdr_wht"])
    for ri, (theme, cnts) in enumerate(themes):
        r = 24 + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_lft"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        for i, ik in enumerate(G1_ISSUES):
            ws.write_formula(r, i + 2,
                '=COUNTIFS(A%d:A1048576,"%s",C%d:C1048576,A%d)' % (T3_XL, ik, T3_XL, r + 1),
                fmts["num"])

    ws.write(31, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 24)
    T3_HDRS = [
        "Error Type", "Address", "Page Theme 1", "Page Theme 2",
        "Indexability", "Indexability Status",
        "HTML hreflang 1", "HTML hreflang 1 URL",
        "HTTP hreflang 1", "HTTP hreflang 1 URL",
        "Sitemap hreflang 1", "Sitemap hreflang 1 URL",
        "Impressions", "Clicks", "Organic Sessions",
    ]
    LFT = {0, 1, 6, 7, 8, 9, 10, 11}
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_lft"] if ci in LFT else fmts["t3_hdr"])
    for ri, row in enumerate(rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_lft"] if ci in LFT else fmts["t3_cell"])
    return themes, G1_ISSUES


# ── SHEET 2 ───────────────────────────────────────────────────────────────────
def build_sheet2(wb, fmts, folder, rulebook, internal_map, gsc_map, ga_map, total_indexable):
    ws = wb.add_worksheet("Hreflang - Combined Group 2")
    for col, w in {"A":31,"B":26,"C":15,"D":16,"E":19,"F":16,"G":21,"H":18,
                   "I":17,"J":26,"K":16,"L":23,"M":15,"N":19,"O":15,
                   "P":15,"Q":17,"R":16,"S":20,"T":13,"U":16}.items():
        ws.set_column("%s:%s" % (col, col), w)

    G2_KEYS = ["hreflang_missing_self_reference", "hreflang_not_using_canonical"]
    G2_DISP = ["Hreflang Missing Self Reference", "Hreflang Not Using Canonical"]

    ws.merge_range(1, 0, 8, 5, "", fmts["summary"])

    rows = []
    df_msr = load_csv(folder, "hreflang_missing_self_reference.csv")
    if not df_msr.empty:
        addr = _gc(df_msr, "Address")
        occ  = _gc(df_msr, "Occurrences")
        idxc = _gc(df_msr, "Indexability")
        idxs = _gc(df_msr, "Indexability Status")
        if addr:
            for _, row in df_msr.iterrows():
                url = _clean(row.get(addr, ""))
                if not url or url == "-":
                    continue
                int_d = internal_map.get(url, {})
                t1, t2, _, pri = _classify(url, rulebook)
                gsc = gsc_map.get(url, {})
                rec = {
                    "_issue_key":    "hreflang_missing_self_reference",
                    "Error Type":    "hreflang_missing_self_reference",
                    "Address":       url,
                    "Occurrences":   safe_num(row.get(occ)) if occ else None,
                    "Page Theme 1":  t1,
                    "Page Theme 2":  t2,
                    "Indexability":        int_d.get("indexability",        _clean(row.get(idxc, "")) if idxc else "-"),
                    "Indexability Status": int_d.get("indexability_status", _clean(row.get(idxs, "")) if idxs else "-"),
                    "Non Canonical Return Link URL": "-",
                    "Canonical":                    "-",
                    "Impressions":      gsc.get("impressions"),
                    "Clicks":           gsc.get("clicks"),
                    "Organic Sessions": ga_map.get(url),
                    "_priority": pri,
                }
                for prefix in ["HTML hreflang", "HTTP hreflang", "Sitemap hreflang"]:
                    n = 1
                    while True:
                        cn  = _gc(df_msr, "%s %d" % (prefix, n))
                        cnu = _gc(df_msr, "%s %d URL" % (prefix, n))
                        if cn is None and cnu is None:
                            break
                        rec["%s %d" % (prefix, n)]     = _clean(row.get(cn,  "")) if cn  else "-"
                        rec["%s %d URL" % (prefix, n)] = _clean(row.get(cnu, "")) if cnu else "-"
                        n += 1
                rows.append(rec)

    df_nuc = load_csv(folder, "hreflang_not_using_canonical.csv")
    if not df_nuc.empty:
        addr = _gc(df_nuc, "URL", "Address")
        ncr  = _gc(df_nuc, "Non Canonical Return Link URL")
        can  = _gc(df_nuc, "Canonical")
        if addr:
            for _, row in df_nuc.iterrows():
                url = _clean(row.get(addr, ""))
                if not url or url == "-":
                    continue
                int_d = internal_map.get(url, {})
                t1, t2, _, pri = _classify(url, rulebook)
                gsc = gsc_map.get(url, {})
                rows.append({
                    "_issue_key":    "hreflang_not_using_canonical",
                    "Error Type":    "hreflang_not_using_canonical",
                    "Address":       url,
                    "Occurrences":   None,
                    "Page Theme 1":  t1,
                    "Page Theme 2":  t2,
                    "Indexability":        int_d.get("indexability",        "-"),
                    "Indexability Status": int_d.get("indexability_status", "-"),
                    "Non Canonical Return Link URL": _clean(row.get(ncr, "")) if ncr else "-",
                    "Canonical":                    _clean(row.get(can, "")) if can else "-",
                    "Impressions":      gsc.get("impressions"),
                    "Clicks":           gsc.get("clicks"),
                    "Organic Sessions": ga_map.get(url),
                    "_priority": pri,
                })

    for ik in G2_KEYS:
        if not any(r["_issue_key"] == ik for r in rows):
            rows.append({
                "_issue_key": ik, "Error Type": ik, "Address": "-",
                "Occurrences": None, "Page Theme 1": "-", "Page Theme 2": "-",
                "Indexability": "-", "Indexability Status": "-",
                "Non Canonical Return Link URL": "-", "Canonical": "-",
                "Impressions": None, "Clicks": None, "Organic Sessions": None,
                "_priority": "N/A",
            })

    rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))
    themes = sorted_themes_from_rows(rows, G2_KEYS)

    # Always include standard hreflang cols; add extras dynamically from data
    std_hreflang = [
        "HTML hreflang 1", "HTML hreflang 1 URL",
        "HTML hreflang 2", "HTML hreflang 2 URL",
        "HTML hreflang 3", "HTML hreflang 3 URL",
        "HTML hreflang 4", "HTML hreflang 4 URL",
        "HTTP hreflang 1", "HTTP hreflang 1 URL",
        "Sitemap hreflang 1", "Sitemap hreflang 1 URL",
    ]
    seen_h = set(std_hreflang)
    extra_cols = []
    for rec in rows:
        for k in rec:
            if k.startswith(("HTML hreflang", "HTTP hreflang", "Sitemap hreflang")) and k not in seen_h:
                extra_cols.append(k)
                seen_h.add(k)
    hreflang_dyn_cols = std_hreflang + extra_cols

    T3_HDRS = (
        ["Error Type", "Address", "Occurrences", "Page Theme 1", "Page Theme 2",
         "Indexability", "Indexability Status"]
        + hreflang_dyn_cols
        + ["Non Canonical Return Link URL", "Canonical",
           "Impressions", "Clicks", "Organic Sessions"]
    )

    PT1_COL_XL = xlsxwriter.utility.xl_col_to_name(3)
    T3_HDR  = 30
    T3_DATA = 31
    T3_XL   = T3_DATA + 1

    ws.write(11, 0, "Table 1", fmts["lbl"])
    ws.set_row(12, 29)
    ws.write(12, 0, "URL Issue Types", fmts["red_lft"])
    for i, d in enumerate(G2_DISP):
        ws.write(12, i + 1, d, fmts["t1_issue_hdr"])
    ws.write(13, 0, "Issue Priority", fmts["red_lft"])
    for i in range(2):
        ws.write(13, i + 1, "High", fmts["cell"])
    ws.write(14, 0, "#Affected URLs", fmts["red_lft"])
    for i, ik in enumerate(G2_KEYS):
        ws.write_formula(14, i + 1,
            '=COUNTIF(A%d:A1048576,"%s")' % (T3_XL, ik), fmts["num"])
    ws.set_row(15, 21)
    ws.write(15, 0, "% Share against Total  HTML URLs Crawled", fmts["red_lft"])
    for i in range(2):
        cl = xlsxwriter.utility.xl_col_to_name(i + 1)
        ws.write_formula(15, i + 1, "=%s15/%d" % (cl, total_indexable), fmts["pct"])

    ws.write(18, 0, "Table 2", fmts["lbl"])
    ws.merge_range(19, 0, 19, 3, "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(20, 29)
    ws.write(20, 0, "Page Theme 1",               fmts["t2_hdr_lft"])
    ws.write(20, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_ctr"])
    for i, d in enumerate(G2_DISP):
        ws.write(20, i + 2, d, fmts["t2_hdr_wht"])
    for ri, (theme, cnts) in enumerate(themes):
        r = 21 + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_lft"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        for i, ik in enumerate(G2_KEYS):
            ws.write_formula(r, i + 2,
                '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,A%d)' % (
                    T3_XL, ik, PT1_COL_XL, T3_XL, PT1_COL_XL, r + 1),
                fmts["num"])

    ws.write(29, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 24)
    url_col_names = {"Non Canonical Return Link URL", "Canonical"}
    url_col_names.update(k for k in T3_HDRS if "URL" in k and "hreflang" in k.lower())
    LFT = {ci for ci, h in enumerate(T3_HDRS) if ci in (0, 1) or h in url_col_names}
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_lft"] if ci in LFT else fmts["t3_hdr"])
    for ri, row in enumerate(rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            elif h == "Occurrences":
                ws.write(r, ci, safe_num(v) if safe_num(v) is not None else "-", fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_lft"] if ci in LFT else fmts["t3_cell"])
    return themes, G2_KEYS, T3_HDR, T3_DATA


# ── SHEET 3 ───────────────────────────────────────────────────────────────────
def build_sheet3(wb, fmts, folder, rulebook, internal_map, gsc_map, ga_map, total_indexable):
    ws = wb.add_worksheet("Hreflang -Combined Group 3")
    for col, w in {"A":27,"B":14,"C":42,"D":16,"E":18,"F":16,"G":14,"H":48,
                   "I":14,"J":20,"K":16,"L":14,"M":14,"N":14,"O":15,"P":15,"Q":15}.items():
        ws.set_column("%s:%s" % (col, col), w)

    G3_KEYS = ["hreflang_non200_hreflang_urls", "hreflang_unlinked_hreflang_urls"]

    ws.merge_range(0, 0, 3, 5, "", fmts["summary"])
    ws.set_row(7, 20)
    ws.merge_range(7, 0, 7, 5, "Summary Table:", fmts["summary"])

    rows = []
    for label, csv_file in [
        ("hreflang_non200_hreflang_urls",   "hreflang_non200_hreflang_urls.csv"),
        ("hreflang_unlinked_hreflang_urls", "hreflang_unlinked_hreflang_urls.csv"),
    ]:
        df = load_csv(folder, csv_file)
        if df.empty:
            continue
        type_c = _gc(df, "Type")
        src_c  = _gc(df, "Source")
        dst_c  = _gc(df, "Destination")
        hre_c  = _gc(df, "hreflang")
        sta_c  = _gc(df, "Status")
        lpos_c = _gc(df, "Link Position")
        lori_c = _gc(df, "Link Origin")
        if not src_c:
            continue
        for _, row in df.iterrows():
            src = _clean(row.get(src_c, ""))
            dst = _clean(row.get(dst_c, "")) if dst_c else "-"
            if not src or src == "-":
                continue
            src_int = internal_map.get(src, {})
            dst_int = internal_map.get(dst, {})
            src_idx = src_int.get("indexability", "-")
            t1, t2, _, pri = _classify(src, rulebook)
            gsc = gsc_map.get(src, {})
            rows.append({
                "_issue_key":          label,
                "_src_indexable":      src_idx.lower() == "indexable",
                "Error Type":          label,
                "Type":                _clean(row.get(type_c, "")) if type_c else "-",
                "Source":              src,
                "Source Status Code":  src_int.get("status_code",  "-"),
                "Source Indexability": src_idx,
                "Page Theme 1":        t1,
                "Page Theme 2":        t2,
                "Destination":               dst,
                "hreflang":                  _clean(row.get(hre_c, "")) if hre_c else "-",
                "Destination Indexability":  dst_int.get("indexability", "-"),
                "Destination Status Code":   dst_int.get("status_code",  "-"),
                "Status":        _clean(row.get(sta_c,  "")) if sta_c  else "-",
                "Link Position": _clean(row.get(lpos_c, "")) if lpos_c else "-",
                "Link Origin":   _clean(row.get(lori_c, "")) if lori_c else "-",
                "Impressions":      gsc.get("impressions"),
                "Clicks":           gsc.get("clicks"),
                "Organic Sessions": ga_map.get(src),
                "_priority":        pri,
            })

    for ik in G3_KEYS:
        if not any(r["_issue_key"] == ik for r in rows):
            rows.append({
                "_issue_key": ik, "_src_indexable": False,
                "Error Type": ik, "Type": "-", "Source": "-",
                "Source Status Code": "-", "Source Indexability": "-",
                "Page Theme 1": "-", "Page Theme 2": "-",
                "Destination": "-", "hreflang": "-",
                "Destination Indexability": "-", "Destination Status Code": "-",
                "Status": "-", "Link Position": "-", "Link Origin": "-",
                "Impressions": None, "Clicks": None, "Organic Sessions": None,
                "_priority": "N/A",
            })

    rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))
    idx_rows = [r for r in rows if r["_src_indexable"]]
    themes   = sorted_themes_from_rows(idx_rows if idx_rows else rows, G3_KEYS)

    T3_HDRS = [
        "Error Type", "Type", "Source", "Source Status Code", "Source Indexability",
        "Page Theme 1", "Page Theme 2", "Destination", "hreflang",
        "Destination Indexability", "Destination Status Code",
        "Status", "Link Position", "Link Origin",
        "Impressions", "Clicks", "Organic Sessions",
    ]
    PT1_COL_XL = xlsxwriter.utility.xl_col_to_name(5)  # F = Page Theme 1
    IDX_COL_XL = xlsxwriter.utility.xl_col_to_name(4)  # E = Source Indexability

    T1_ROW   = 11
    T2_ROW   = T1_ROW + 7
    T2_TITLE = T2_ROW + 1
    T2_COLHD = T2_ROW + 2
    T2_DATA0 = T2_ROW + 3
    T2_DATAN = T2_DATA0 + len(themes) - 1
    T3_LBL   = T2_DATAN + 2
    T3_HDR   = T3_LBL + 1
    T3_DATA  = T3_HDR + 1
    T3_XL    = T3_DATA + 1

    ws.write(T1_ROW, 0, "Table 1", fmts["lbl"])
    ws.set_row(T1_ROW + 1, 29)
    ws.write(T1_ROW + 1, 0, "Error Type", fmts["red_lft"])
    for i, d in enumerate(G3_KEYS):
        ws.write(T1_ROW + 1, i + 1, d, fmts["t1_issue_hdr"])
    ws.write(T1_ROW + 2, 0, "Issue Priority", fmts["red_lft"])
    for i in range(2):
        ws.write(T1_ROW + 2, i + 1, "Medium" if i == 0 else "High", fmts["cell"])
    ws.write(T1_ROW + 3, 0, "#Affected URLs", fmts["red_lft"])
    for i, ik in enumerate(G3_KEYS):
        ws.write_formula(T1_ROW + 3, i + 1,
            '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable")' % (
                T3_XL, ik, IDX_COL_XL, T3_XL, IDX_COL_XL),
            fmts["num"])
    ws.set_row(T1_ROW + 4, 21)
    ws.write(T1_ROW + 4, 0, "% Share against Total  Indexable HTML URLs Crawled", fmts["red_lft"])
    for i in range(2):
        cl = xlsxwriter.utility.xl_col_to_name(i + 1)
        ws.write_formula(T1_ROW + 4, i + 1,
            "=%s%d/%d" % (cl, T1_ROW + 4, total_indexable), fmts["pct"])

    ws.write(T2_ROW, 0, "Table 2", fmts["lbl"])
    ws.merge_range(T2_TITLE, 0, T2_TITLE, 3, "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(T2_COLHD, 29)
    ws.write(T2_COLHD, 0, "Page Theme 1",               fmts["t2_hdr_lft"])
    ws.write(T2_COLHD, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_ctr"])
    for i, d in enumerate(G3_KEYS):
        ws.write(T2_COLHD, i + 2, d, fmts["t2_hdr_wht"])
    for ri, (theme, cnts) in enumerate(themes):
        r = T2_DATA0 + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_lft"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        for i, ik in enumerate(G3_KEYS):
            ws.write_formula(r, i + 2,
                '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable",%s%d:%s1048576,A%d)' % (
                    T3_XL, ik, IDX_COL_XL, T3_XL, IDX_COL_XL,
                    PT1_COL_XL, T3_XL, PT1_COL_XL, r + 1),
                fmts["num"])

    ws.write(T3_LBL, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 24)
    LFT = {ci for ci, h in enumerate(T3_HDRS)
           if ci in (0, 1, 2, 7) or "URL" in h or h in ("Source", "Destination")}
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_lft"] if ci in LFT else fmts["t3_hdr"])
    for ri, row in enumerate(rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_lft"] if ci in LFT else fmts["t3_cell"])
    return T3_HDR, T3_DATA


# ── SHEET 4 ───────────────────────────────────────────────────────────────────
def build_sheet4(wb, fmts, folder, rulebook, internal_map, gsc_map, ga_map, total_indexable):
    ws = wb.add_worksheet("hreflang_inconsistent_language_")
    for col, w in {"A":30,"B":32,"C":20,"D":22,"E":18,"F":18,
                   "G":28,"H":32,"I":18,"J":16,"K":15,"L":15,"M":15}.items():
        ws.set_column("%s:%s" % (col, col), w)

    KEY = "hreflang_inconsistent_language_return_links"
    ws.merge_range(0, 0, 4, 5, "", fmts["summary"])
    ws.set_row(8, 20)
    ws.merge_range(8, 0, 8, 5, "Summary Table:", fmts["summary"])

    df = load_csv(folder, "hreflang_inconsistent_language_region_return_links.csv")
    if df.empty:
        df = load_csv(folder, "hreflang_inconsistent_language_return_links.csv")

    url_col = _gc(df, "URL with Inconsistent Language Return Link", "Address", "URL")
    ret_col = _gc(df, "URL Returning with Inconsistent Language")
    exp_col = _gc(df, "Expected Language")
    act_col = _gc(df, "Actual Language")
    tgt_col = _gc(df, "URL Target")

    rows = []
    if not df.empty and url_col:
        for _, row in df.iterrows():
            url = _clean(row.get(url_col, ""))
            if not url or url == "-":
                continue
            int_d = internal_map.get(url, {})
            t1, t2, _, pri = _classify(url, rulebook)
            gsc = gsc_map.get(url, {})
            rows.append({
                "_issue_key":  KEY,
                "_indexable":  int_d.get("indexability", "-").lower() == "indexable",
                "Error Type":  KEY,
                "URL with Inconsistent Language Return Link":               url,
                "URL with Inconsistent Language Return Link - Status Code":  int_d.get("status_code",  "-"),
                "URL with Inconsistent Language Return Link - Indexability": int_d.get("indexability", "-"),
                "Page Theme 1":  t1,
                "Page Theme 2":  t2,
                "URL Target":    _clean(row.get(tgt_col, "")) if tgt_col else "-",
                "URL Returning with Inconsistent Language": _clean(row.get(ret_col, "")) if ret_col else "-",
                "Expected Language": _clean(row.get(exp_col, "")) if exp_col else "-",
                "Actual Language":   _clean(row.get(act_col, "")) if act_col else "-",
                "Impressions":      gsc.get("impressions"),
                "Clicks":           gsc.get("clicks"),
                "Organic Sessions": ga_map.get(url),
                "_priority": pri,
            })

    if not rows:
        rows.append({
            "_issue_key": KEY, "_indexable": False, "Error Type": KEY,
            "URL with Inconsistent Language Return Link": "-",
            "URL with Inconsistent Language Return Link - Status Code": "-",
            "URL with Inconsistent Language Return Link - Indexability": "-",
            "Page Theme 1": "-", "Page Theme 2": "-",
            "URL Target": "-", "URL Returning with Inconsistent Language": "-",
            "Expected Language": "-", "Actual Language": "-",
            "Impressions": None, "Clicks": None, "Organic Sessions": None,
            "_priority": "N/A",
        })

    rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))
    idx_rows = [r for r in rows if r["_indexable"]]
    themes   = sorted_themes_from_rows(idx_rows if idx_rows else rows, [KEY])

    T3_HDRS = [
        "Error Type",
        "URL with Inconsistent Language Return Link",
        "URL with Inconsistent Language Return Link - Status Code",
        "URL with Inconsistent Language Return Link - Indexability",
        "Page Theme 1", "Page Theme 2",
        "URL Target",
        "URL Returning with Inconsistent Language",
        "Expected Language", "Actual Language",
        "Impressions", "Clicks", "Organic Sessions",
    ]
    PT1_COL_XL = xlsxwriter.utility.xl_col_to_name(4)  # E
    IDX_COL_XL = xlsxwriter.utility.xl_col_to_name(3)  # D

    T2_HDR_ROW    = 16
    T2_TITLE      = T2_HDR_ROW + 1
    T2_COL_HDR    = T2_HDR_ROW + 2
    T2_DATA_START = T2_HDR_ROW + 3
    T2_DATA_END   = T2_DATA_START + len(themes) - 1
    T3_LBL  = T2_DATA_END + 2
    T3_HDR  = T3_LBL + 1
    T3_DATA = T3_HDR + 1
    T3_XL   = T3_DATA + 1

    ws.write(10, 0, "Table 1", fmts["lbl"])
    ws.set_row(11, 29)
    ws.write(11, 0, "URL Issue Types", fmts["red_lft"])
    ws.write(11, 1, "URL with Inconsistent Language Return Link", fmts["t1_issue_hdr"])
    ws.write(12, 0, "Issue Priority", fmts["red_lft"])
    ws.write(12, 1, "High", fmts["cell"])
    ws.write(13, 0, "#Affected URLs", fmts["red_lft"])
    ws.write_formula(13, 1,
        '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable")' % (
            T3_XL, KEY, IDX_COL_XL, T3_XL, IDX_COL_XL),
        fmts["num"])
    ws.set_row(14, 21)
    ws.write(14, 0, "% Share against Total  URLs Crawled", fmts["red_lft"])
    ws.write_formula(14, 1, "=B14/%d" % total_indexable, fmts["pct"])

    ws.write(T2_HDR_ROW, 0, "Table 2", fmts["lbl"])
    ws.merge_range(T2_TITLE, 0, T2_TITLE, 2, "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(T2_COL_HDR, 29)
    ws.write(T2_COL_HDR, 0, "Page Theme 1",               fmts["t2_hdr_lft"])
    ws.write(T2_COL_HDR, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_ctr"])
    ws.write(T2_COL_HDR, 2, "URL with Inconsistent Language Return Link", fmts["t2_hdr_wht"])
    for ri, (theme, cnts) in enumerate(themes):
        r = T2_DATA_START + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_lft"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        ws.write_formula(r, 2,
            '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable",%s%d:%s1048576,A%d)' % (
                T3_XL, KEY,
                IDX_COL_XL, T3_XL, IDX_COL_XL,
                PT1_COL_XL, T3_XL, PT1_COL_XL, r + 1),
            fmts["num"])

    ws.write(T3_LBL, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 30)
    LFT = {0, 1, 6, 7}
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_lft"] if ci in LFT else fmts["t3_hdr"])
    for ri, row in enumerate(rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_lft"] if ci in LFT else fmts["t3_cell"])
    return T3_HDR, T3_DATA


# ── SHEET 5 ───────────────────────────────────────────────────────────────────
def build_sheet5(wb, fmts, folder, rulebook, internal_map, gsc_map, ga_map, total_indexable):
    ws = wb.add_worksheet("Hreflang missing Return links")
    for col, w in {"A":28,"B":30,"C":20,"D":22,"E":18,"F":18,
                   "G":30,"H":20,"I":22,"J":30,"K":20,"L":22,"M":15,"N":15,"O":15}.items():
        ws.set_column("%s:%s" % (col, col), w)

    KEY  = "hreflang_missing_return_links"
    DISP = "Hreflang missing Return links"
    ws.merge_range(0, 0, 8, 5, "", fmts["summary"])
    ws.set_row(11, 20)
    ws.merge_range(11, 0, 11, 5, "Summary Table:", fmts["summary"])

    df = load_csv(folder, "hreflang_missing_return_links.csv")
    url_col = _gc(df, "URL Missing Return Link", "Address")
    nrl_col = _gc(df, "URL Not Returning Link")
    exp_col = _gc(df, "Expected Link")

    rows = []
    if not df.empty and url_col:
        for _, row in df.iterrows():
            url = _clean(row.get(url_col, ""))
            if not url or url == "-":
                continue
            nrl = _clean(row.get(nrl_col, "")) if nrl_col else "-"
            exp = _clean(row.get(exp_col, "")) if exp_col else "-"
            int_url = internal_map.get(url, {})
            int_nrl = internal_map.get(nrl, {})
            int_exp = internal_map.get(exp, {})
            t1, t2, _, pri = _classify(url, rulebook)
            gsc = gsc_map.get(url, {})
            rows.append({
                "_issue_key": KEY,
                "_indexable": int_url.get("indexability", "-").lower() == "indexable",
                "Error Type": KEY,
                "URL Missing Return Link":               url,
                "URL Missing Return Link- Status Code":  int_url.get("status_code",  "-"),
                "URL Missing Return Link- Indexability": int_url.get("indexability", "-"),
                "Page Theme 1":  t1,
                "Page Theme 2":  t2,
                "URL Not Returning Link":                nrl,
                "URL Not Returning Link - Status Code":  int_nrl.get("status_code",  "-"),
                "URL Not Returning Link - Indexability": int_nrl.get("indexability", "-"),
                "Expected Link":                         exp,
                "Expected Link - Status Code":           int_exp.get("status_code",  "-"),
                "Expected Link-  Indexability":          int_exp.get("indexability", "-"),
                "Impressions":      gsc.get("impressions"),
                "Clicks":           gsc.get("clicks"),
                "Organic Sessions": ga_map.get(url),
                "_priority": pri,
            })

    if not rows:
        rows.append({
            "_issue_key": KEY, "_indexable": False, "Error Type": KEY,
            "URL Missing Return Link": "-", "URL Missing Return Link- Status Code": "-",
            "URL Missing Return Link- Indexability": "-",
            "Page Theme 1": "-", "Page Theme 2": "-",
            "URL Not Returning Link": "-", "URL Not Returning Link - Status Code": "-",
            "URL Not Returning Link - Indexability": "-",
            "Expected Link": "-", "Expected Link - Status Code": "-",
            "Expected Link-  Indexability": "-",
            "Impressions": None, "Clicks": None, "Organic Sessions": None, "_priority": "N/A",
        })

    rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))
    idx_rows = [r for r in rows if r["_indexable"]]
    themes   = sorted_themes_from_rows(idx_rows if idx_rows else rows, [KEY])

    T3_HDRS = [
        "Error Type",
        "URL Missing Return Link", "URL Missing Return Link- Status Code",
        "URL Missing Return Link- Indexability",
        "Page Theme 1", "Page Theme 2",
        "URL Not Returning Link", "URL Not Returning Link - Status Code",
        "URL Not Returning Link - Indexability",
        "Expected Link", "Expected Link - Status Code", "Expected Link-  Indexability",
        "Impressions", "Clicks", "Organic Sessions",
    ]
    IDX_COL_XL = xlsxwriter.utility.xl_col_to_name(3)  # D
    PT1_COL_XL = xlsxwriter.utility.xl_col_to_name(4)  # E

    T2_HDR_ROW    = 20
    T2_TITLE      = T2_HDR_ROW + 1
    T2_COL_HDR    = T2_HDR_ROW + 2
    T2_DATA_START = T2_HDR_ROW + 3
    T2_DATA_END   = T2_DATA_START + len(themes) - 1
    T3_LBL  = T2_DATA_END + 2
    T3_HDR  = T3_LBL + 1
    T3_DATA = T3_HDR + 1
    T3_XL   = T3_DATA + 1

    ws.write(13, 0, "Table 1", fmts["lbl"])
    ws.set_row(14, 29)
    ws.write(14, 0, "URL Issue Types", fmts["red_lft"])
    ws.write(14, 1, DISP, fmts["t1_issue_hdr"])
    ws.write(15, 0, "Issue Priority", fmts["red_lft"])
    ws.write(15, 1, "High", fmts["cell"])
    ws.write(16, 0, "#Affected URLs", fmts["red_lft"])
    ws.write_formula(16, 1,
        '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable")' % (
            T3_XL, KEY, IDX_COL_XL, T3_XL, IDX_COL_XL),
        fmts["num"])
    ws.set_row(17, 21)
    ws.write(17, 0, "% Share against Total  URLs Crawled", fmts["red_lft"])
    ws.write_formula(17, 1, "=B17/%d" % total_indexable, fmts["pct"])

    ws.write(T2_HDR_ROW, 0, "Table 2", fmts["lbl"])
    ws.merge_range(T2_TITLE, 0, T2_TITLE, 2, "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(T2_COL_HDR, 29)
    ws.write(T2_COL_HDR, 0, "Page Theme 1",               fmts["t2_hdr_lft"])
    ws.write(T2_COL_HDR, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_ctr"])
    ws.write(T2_COL_HDR, 2, "Total Pages missing Return links", fmts["t2_hdr_wht"])
    for ri, (theme, cnts) in enumerate(themes):
        r = T2_DATA_START + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_lft"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        ws.write_formula(r, 2,
            '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable",%s%d:%s1048576,A%d)' % (
                T3_XL, KEY,
                IDX_COL_XL, T3_XL, IDX_COL_XL,
                PT1_COL_XL, T3_XL, PT1_COL_XL, r + 1),
            fmts["num"])

    ws.write(T3_LBL, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 42)
    LFT = {0, 1, 6, 9}
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_lft"] if ci in LFT else fmts["t3_hdr"])
    for ri, row in enumerate(rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_lft"] if ci in LFT else fmts["t3_cell"])
    return T3_HDR, T3_DATA


# ── SHEET 6 ───────────────────────────────────────────────────────────────────
def build_sheet6(wb, fmts, folder, rulebook, internal_map, gsc_map, ga_map, total_indexable):
    ws = wb.add_worksheet("hreflang_no_index_return_links")
    for col, w in {"A":28,"B":30,"C":16,"D":18,"E":18,"F":18,
                   "G":30,"H":18,"I":20,"J":20,"K":20,"L":14,"M":15,"N":15,"O":15}.items():
        ws.set_column("%s:%s" % (col, col), w)

    KEY  = "hreflang_no_index_return_links"
    DISP = "hreflang_no_index_return_links"
    ws.merge_range(0, 0, 9, 5, "", fmts["summary"])
    ws.set_row(13, 20)
    ws.merge_range(13, 0, 13, 5, "Summary Table:", fmts["summary"])

    df = load_csv(folder, "hreflang_noindex_return_links.csv")
    if df.empty:
        df = load_csv(folder, "hreflang_no_index_return_links.csv")

    url_col  = _gc(df, "URL", "Address")
    ni_col   = _gc(df, "Noindex URL", "No index URL", "Noindex_URL")
    lang_col = _gc(df, "Language")

    rows = []
    if not df.empty and url_col:
        for _, row in df.iterrows():
            url = _clean(row.get(url_col, ""))
            if not url or url == "-":
                continue
            ni_url = _clean(row.get(ni_col, "")) if ni_col else "-"
            int_url = internal_map.get(url, {})
            int_ni  = internal_map.get(ni_url, {})
            t1,  t2,  _, pri = _classify(url,    rulebook)
            ni1, ni2, _, _   = _classify(ni_url, rulebook)
            gsc = gsc_map.get(url, {})
            rows.append({
                "_issue_key":          KEY,
                "_indexable":          int_url.get("indexability", "-").lower() == "indexable",
                "Error Type":          KEY,
                "URL":                 url,
                "URL Status Code":     int_url.get("status_code",  "-"),
                "URL - Indexability":  int_url.get("indexability", "-"),
                "Page Theme 1":        t1,
                "Page Theme 2":        t2,
                "Noindex URL":                ni_url,
                "No index URL- Status Code":  int_ni.get("status_code",  "-"),
                "Noindex URL - Indexability": int_ni.get("indexability", "-"),
                "Noindex URL - Page Theme 1": ni1,
                "Noindex URL - Page Theme 2": ni2,
                "Language":            _clean(row.get(lang_col, "")) if lang_col else "-",
                "Impressions":         gsc.get("impressions"),
                "Clicks":              gsc.get("clicks"),
                "Organic Sessions":    ga_map.get(url),
                "_priority":           pri,
            })

    if not rows:
        rows.append({
            "_issue_key": KEY, "_indexable": False, "Error Type": KEY, "URL": "-",
            "URL Status Code": "-", "URL - Indexability": "-",
            "Page Theme 1": "-", "Page Theme 2": "-",
            "Noindex URL": "-", "No index URL- Status Code": "-",
            "Noindex URL - Indexability": "-",
            "Noindex URL - Page Theme 1": "-", "Noindex URL - Page Theme 2": "-",
            "Language": "-",
            "Impressions": None, "Clicks": None, "Organic Sessions": None, "_priority": "N/A",
        })

    rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))
    idx_rows = [r for r in rows if r["_indexable"]]
    themes   = sorted_themes_from_rows(idx_rows if idx_rows else rows, [KEY])

    T3_HDRS = [
        "Error Type", "URL", "URL Status Code", "URL - Indexability",
        "Page Theme 1", "Page Theme 2",
        "Noindex URL", "No index URL- Status Code", "Noindex URL - Indexability",
        "Noindex URL - Page Theme 1", "Noindex URL - Page Theme 2",
        "Language", "Impressions", "Clicks", "Organic Sessions",
    ]
    IDX_COL_XL = xlsxwriter.utility.xl_col_to_name(3)  # D
    PT1_COL_XL = xlsxwriter.utility.xl_col_to_name(4)  # E

    T2_TITLE = 21
    T2_COLHD = 22
    T2_DATA0 = 23
    T2_DATAN = T2_DATA0 + len(themes) - 1
    T3_LBL   = T2_DATAN + 2
    T3_HDR   = T3_LBL + 1
    T3_DATA  = T3_HDR + 1
    T3_XL    = T3_DATA + 1

    ws.write(15, 0, "Table 1", fmts["lbl"])
    ws.set_row(16, 29)
    ws.write(16, 0, "URL Issue Types", fmts["red_lft"])
    ws.write(16, 1, DISP, fmts["t1_issue_hdr"])
    ws.write(17, 0, "Issue Priority", fmts["red_lft"])
    ws.write(17, 1, "High", fmts["cell"])
    ws.write(18, 0, "#Affected URLs", fmts["red_lft"])
    ws.write_formula(18, 1,
        '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable")' % (
            T3_XL, KEY, IDX_COL_XL, T3_XL, IDX_COL_XL),
        fmts["num"])
    ws.set_row(19, 21)
    ws.write(19, 0, "% Share against Total HTML URLs Crawled", fmts["red_lft"])
    ws.write_formula(19, 1, "=B19/%d" % total_indexable, fmts["pct"])

    ws.merge_range(T2_TITLE, 0, T2_TITLE, 2, "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(T2_COLHD, 29)
    ws.write(T2_COLHD, 0, "Page Theme 1",               fmts["t2_hdr_lft"])
    ws.write(T2_COLHD, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_ctr"])
    ws.write(T2_COLHD, 2, DISP,                          fmts["t2_hdr_wht"])
    for ri, (theme, cnts) in enumerate(themes):
        r = T2_DATA0 + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_lft"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        ws.write_formula(r, 2,
            '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable",%s%d:%s1048576,A%d)' % (
                T3_XL, KEY,
                IDX_COL_XL, T3_XL, IDX_COL_XL,
                PT1_COL_XL, T3_XL, PT1_COL_XL, r + 1),
            fmts["num"])

    ws.write(T3_LBL, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 30)
    LFT = {0, 1, 6}
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_lft"] if ci in LFT else fmts["t3_hdr"])
    for ri, row in enumerate(rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_lft"] if ci in LFT else fmts["t3_cell"])
    return T3_HDR, T3_DATA


# ── SHEET 7 ───────────────────────────────────────────────────────────────────
def build_sheet7(wb, fmts, folder, rulebook, internal_map, gsc_map, ga_map, total_indexable):
    ws = wb.add_worksheet("hreflang_multiple_entries")
    for col, w in {"A":22,"B":22,"C":18,"D":20,"E":18,"F":18,
                   "G":22,"H":22,"I":22,"J":22,"K":22,"L":22,
                   "M":22,"N":22,"O":22,"P":22,"Q":15,"R":15,"S":15}.items():
        ws.set_column("%s:%s" % (col, col), w)

    KEY  = "hreflang_multiple_entries"
    DISP = "Hreflang Multiple Entries"
    ws.merge_range(0, 0, 7, 5, "", fmts["summary"])
    ws.set_row(11, 20)
    ws.merge_range(11, 0, 11, 5, "Summary Table:", fmts["summary"])

    df = load_csv(folder, "hreflang_multiple_entries.csv")
    addr_c = _gc(df, "Address")
    idx_c  = _gc(df, "Indexability")
    idxs_c = _gc(df, "Indexability Status")

    rows = []
    if not df.empty and addr_c:
        for _, row in df.iterrows():
            url = _clean(row.get(addr_c, ""))
            if not url or url == "-":
                continue
            int_d = internal_map.get(url, {})
            idx_val  = int_d.get("indexability",        _clean(row.get(idx_c,  "")) if idx_c  else "-")
            idxs_val = int_d.get("indexability_status", _clean(row.get(idxs_c, "")) if idxs_c else "-")
            t1, t2, _, pri = _classify(url, rulebook)
            gsc = gsc_map.get(url, {})
            rec = {
                "_issue_key":  KEY,
                "_indexable":  idx_val.lower() == "indexable",
                "Error Type":  KEY,
                "Address":     url,
                "Indexability":        idx_val,
                "Indexability Status": idxs_val,
                "Page Theme 1": t1,
                "Page Theme 2": t2,
                "Impressions":      gsc.get("impressions"),
                "Clicks":           gsc.get("clicks"),
                "Organic Sessions": ga_map.get(url),
                "_priority": pri,
            }
            for prefix in ["HTML hreflang", "HTTP hreflang", "Sitemap hreflang"]:
                n = 1
                while True:
                    cn  = _gc(df, "%s %d" % (prefix, n))
                    cnu = _gc(df, "%s %d URL" % (prefix, n))
                    if cn is None and cnu is None:
                        break
                    rec["%s %d" % (prefix, n)]     = _clean(row.get(cn,  "")) if cn  else "-"
                    rec["%s %d URL" % (prefix, n)] = _clean(row.get(cnu, "")) if cnu else "-"
                    n += 1
            rows.append(rec)

    if not rows:
        rows.append({
            "_issue_key": KEY, "_indexable": False, "Error Type": KEY, "Address": "-",
            "Indexability": "-", "Indexability Status": "-",
            "Page Theme 1": "-", "Page Theme 2": "-",
            "Impressions": None, "Clicks": None, "Organic Sessions": None, "_priority": "N/A",
        })

    rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))
    idx_rows = [r for r in rows if r["_indexable"]]
    themes   = sorted_themes_from_rows(idx_rows if idx_rows else rows, [KEY])

    fixed_front = ["Error Type", "Address", "Indexability", "Indexability Status",
                   "Page Theme 1", "Page Theme 2"]
    hreflang_dyn = []
    seen_h = set()
    for rec in rows:
        for k in rec:
            if k.startswith(("HTML hreflang", "HTTP hreflang", "Sitemap hreflang")) and k not in seen_h:
                hreflang_dyn.append(k)
                seen_h.add(k)
    T3_HDRS = fixed_front + hreflang_dyn + ["Impressions", "Clicks", "Organic Sessions"]

    IDX_COL_XL = xlsxwriter.utility.xl_col_to_name(2)  # C
    PT1_COL_XL = xlsxwriter.utility.xl_col_to_name(4)  # E

    T2_HDR_ROW = 19
    T2_TITLE   = T2_HDR_ROW + 1
    T2_COLHD   = T2_HDR_ROW + 2
    T2_DATA0   = T2_HDR_ROW + 3
    T2_DATAN   = T2_DATA0 + len(themes) - 1
    T3_LBL     = T2_DATAN + 2
    T3_HDR     = T3_LBL + 1
    T3_DATA    = T3_HDR + 1
    T3_XL      = T3_DATA + 1

    ws.write(13, 0, "Table 1", fmts["lbl"])
    ws.set_row(14, 29)
    ws.write(14, 0, "URL Issue Types", fmts["red_lft"])
    ws.write(14, 1, DISP, fmts["t1_issue_hdr"])
    ws.write(15, 0, "Issue Priority", fmts["red_lft"])
    ws.write(15, 1, "High", fmts["cell"])
    ws.write(16, 0, "#Affected URLs", fmts["red_lft"])
    ws.write_formula(16, 1,
        '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable")' % (
            T3_XL, KEY, IDX_COL_XL, T3_XL, IDX_COL_XL),
        fmts["num"])
    ws.set_row(17, 21)
    ws.write(17, 0, "% Share against Total  HTML URLs Crawled", fmts["red_lft"])
    ws.write_formula(17, 1, "=B17/%d" % total_indexable, fmts["pct"])

    ws.write(T2_HDR_ROW, 0, "Table 2", fmts["lbl"])
    ws.merge_range(T2_TITLE, 0, T2_TITLE, 2, "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(T2_COLHD, 29)
    ws.write(T2_COLHD, 0, "Page Theme 1",               fmts["t2_hdr_lft"])
    ws.write(T2_COLHD, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_ctr"])
    ws.write(T2_COLHD, 2, DISP,                          fmts["t2_hdr_wht"])
    for ri, (theme, cnts) in enumerate(themes):
        r = T2_DATA0 + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_lft"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        ws.write_formula(r, 2,
            '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable",%s%d:%s1048576,A%d)' % (
                T3_XL, KEY,
                IDX_COL_XL, T3_XL, IDX_COL_XL,
                PT1_COL_XL, T3_XL, PT1_COL_XL, r + 1),
            fmts["num"])

    ws.write(T3_LBL, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 30)
    LFT = {0, 1}
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_lft"] if ci in LFT else fmts["t3_hdr"])
    for ri, row in enumerate(rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_lft"] if ci in LFT else fmts["t3_cell"])
    return T3_HDR, T3_DATA


# ── SHEET 8 ───────────────────────────────────────────────────────────────────
def build_sheet8(wb, fmts, folder, rulebook, internal_map, gsc_map, ga_map, total_indexable):
    ws = wb.add_worksheet("hreflang_non_canonical_return_l")
    for col, w in {"A":32,"B":28,"C":18,"D":20,"E":18,"F":18,
                   "G":32,"H":22,"I":24,"J":28,"K":18,"L":20,"M":15,"N":15,"O":15}.items():
        ws.set_column("%s:%s" % (col, col), w)

    KEY  = "hreflang_noncanonical_return_links"
    DISP = "Hreflang Non Canonical Return Link"
    ws.merge_range(2, 0, 8, 5, "", fmts["summary"])
    ws.set_row(12, 20)
    ws.merge_range(12, 0, 12, 5, "Summary Table:", fmts["summary"])

    df = load_csv(folder, "hreflang_noncanonical_return_links.csv")
    if df.empty:
        df = load_csv(folder, "hreflang_not_using_canonical.csv")

    url_col = _gc(df, "URL", "Address")
    ncr_col = _gc(df, "Non Canonical Return Link URL", "Non Canonical Return Link")
    can_col = _gc(df, "Canonical")

    rows = []
    if not df.empty and url_col:
        for _, row in df.iterrows():
            url = _clean(row.get(url_col, ""))
            if not url or url == "-":
                continue
            ncr = _clean(row.get(ncr_col, "")) if ncr_col else "-"
            can = _clean(row.get(can_col, "")) if can_col else "-"
            int_url = internal_map.get(url, {})
            int_ncr = internal_map.get(ncr, {})
            int_can = internal_map.get(can, {})
            t1, t2, _, pri = _classify(url, rulebook)
            gsc = gsc_map.get(url, {})
            rows.append({
                "_issue_key":  KEY,
                "_indexable":  int_url.get("indexability", "-").lower() == "indexable",
                "Error Type":  KEY,
                "URL":         url,
                "URL- Status Code":   int_url.get("status_code",  "-"),
                "URL- Indexability":  int_url.get("indexability", "-"),
                "Page Theme 1": t1,
                "Page Theme 2": t2,
                "Non Canonical Return Link URL":                ncr,
                "Non Canonical Return Link URL - Status Code":  int_ncr.get("status_code",  "-"),
                "Non Canonical Return Link URL - Indexability": int_ncr.get("indexability", "-"),
                "Canonical":                can,
                "Canonical - Status Code":  int_can.get("status_code",  "-"),
                "Canonical - Indexability": int_can.get("indexability", "-"),
                "Impressions":      gsc.get("impressions"),
                "Clicks":           gsc.get("clicks"),
                "Organic Sessions": ga_map.get(url),
                "_priority": pri,
            })

    if not rows:
        rows.append({
            "_issue_key": KEY, "_indexable": False, "Error Type": KEY, "URL": "-",
            "URL- Status Code": "-", "URL- Indexability": "-",
            "Page Theme 1": "-", "Page Theme 2": "-",
            "Non Canonical Return Link URL": "-",
            "Non Canonical Return Link URL - Status Code": "-",
            "Non Canonical Return Link URL - Indexability": "-",
            "Canonical": "-", "Canonical - Status Code": "-", "Canonical - Indexability": "-",
            "Impressions": None, "Clicks": None, "Organic Sessions": None, "_priority": "N/A",
        })

    rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))
    idx_rows = [r for r in rows if r["_indexable"]]
    themes   = sorted_themes_from_rows(idx_rows if idx_rows else rows, [KEY])

    T3_HDRS = [
        "Error Type", "URL", "URL- Status Code", "URL- Indexability",
        "Page Theme 1", "Page Theme 2",
        "Non Canonical Return Link URL",
        "Non Canonical Return Link URL - Status Code",
        "Non Canonical Return Link URL - Indexability",
        "Canonical", "Canonical - Status Code", "Canonical - Indexability",
        "Impressions", "Clicks", "Organic Sessions",
    ]
    IDX_COL_XL = xlsxwriter.utility.xl_col_to_name(3)  # D
    PT1_COL_XL = xlsxwriter.utility.xl_col_to_name(4)  # E

    T2_HDR_ROW = 23
    T2_TITLE   = T2_HDR_ROW + 1
    T2_COLHD   = T2_HDR_ROW + 2
    T2_DATA0   = T2_HDR_ROW + 3
    T2_DATAN   = T2_DATA0 + len(themes) - 1
    T3_LBL     = T2_DATAN + 2
    T3_HDR     = T3_LBL + 1
    T3_DATA    = T3_HDR + 1
    T3_XL      = T3_DATA + 1

    ws.write(16, 0, "Table 1", fmts["lbl"])
    ws.set_row(17, 29)
    ws.write(17, 0, "URL Issue Types", fmts["red_lft"])
    ws.write(17, 1, DISP, fmts["t1_issue_hdr"])
    ws.write(18, 0, "Issue Priority", fmts["red_lft"])
    ws.write(18, 1, "High", fmts["cell"])
    ws.write(19, 0, "#Affected URLs", fmts["red_lft"])
    ws.write_formula(19, 1,
        '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable")' % (
            T3_XL, KEY, IDX_COL_XL, T3_XL, IDX_COL_XL),
        fmts["num"])
    ws.set_row(20, 21)
    ws.write(20, 0, "% Share against Total  URLs Crawled", fmts["red_lft"])
    ws.write_formula(20, 1, "=B20/%d" % total_indexable, fmts["pct"])

    ws.write(T2_HDR_ROW, 0, "Table 2", fmts["lbl"])
    ws.merge_range(T2_TITLE, 0, T2_TITLE, 2, "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(T2_COLHD, 29)
    ws.write(T2_COLHD, 0, "Page Theme 1",               fmts["t2_hdr_lft"])
    ws.write(T2_COLHD, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_ctr"])
    ws.write(T2_COLHD, 2, "Total Pages Non Canonical Return links", fmts["t2_hdr_wht"])
    for ri, (theme, cnts) in enumerate(themes):
        r = T2_DATA0 + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_lft"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        ws.write_formula(r, 2,
            '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,"Indexable",%s%d:%s1048576,A%d)' % (
                T3_XL, KEY,
                IDX_COL_XL, T3_XL, IDX_COL_XL,
                PT1_COL_XL, T3_XL, PT1_COL_XL, r + 1),
            fmts["num"])

    ws.write(T3_LBL, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 43)
    LFT = {0, 1, 6, 9}
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_lft"] if ci in LFT else fmts["t3_hdr"])
    for ri, row in enumerate(rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_lft"] if ci in LFT else fmts["t3_cell"])
    return T3_HDR, T3_DATA


# ── OPENPYXL POST-PROCESSING ──────────────────────────────────────────────────
SUMMARY_PAIRS = {
    0: [
        (True,  "Issue Summary\n"),
        (False, "In the following report, we covered the below hreflang issues:\n-> "),
        (True,  "hreflang_outside_head"),
        (False, " \u2013 URLs with hreflang annotations implemented outside the HTML <head> section.\n-> "),
        (True,  "hreflang_missing"),
        (False, " \u2013 URLs without hreflang annotations.\n-> "),
        (True,  "hreflang_missing_xdefault"),
        (False, " \u2013 URLs missing the x-default hreflang annotation.\n-> "),
        (True,  "hreflang_incorrect_language_region_codes"),
        (False, " \u2013 URLs using invalid or incorrect hreflang language or region codes.\n\n"
                "These issues can affect how search engines discover, interpret, and validate "
                "hreflang implementation across alternate pages."),
    ],
    1: [
        (True,  "Issue Summary\n"),
        (False, "-> "),
        (True,  "hreflang_missing_self_reference"),
        (False, " \u2013 URLs missing a hreflang tag that references the page itself.\n-> "),
        (True,  "hreflang_not_using_canonical"),
        (False, " \u2013 URLs where hreflang tags point to non-canonical URL versions.\n\n"
                "These issues can create inconsistent hreflang signals and prevent search engines "
                "from properly validating alternate page relationships."),
    ],
    2: [
        (True,  "Issue Summary\n\n"),
        (False, "In the following report, we covered the below hreflang issues:\n-> "),
        (True,  "hreflang_non200_hreflang_urls"),
        (False, " \u2013 URLs referenced within rel=\"alternate\" hreflang annotations that do not "
                "return a 200 status code.\n-> "),
        (True,  "hreflang_unlinked_hreflang_urls"),
        (False, " \u2013 URLs that are only discoverable through hreflang annotations and are not "
                "internally linked within the website structure."),
    ],
    3: [
        (True,  "Issue Summary\n"),
        (False, "In the following report, we covered the "),
        (True,  "hreflang_inconsistent_language_return_links"),
        (False, " issue:\nURLs with inconsistent language and regional return links to them."),
    ],
    4: [
        (True,  "Issue Summary\n\n"),
        (False, "In the following report, we covered the Hreflang missing Return links issue:\n\n"),
        (True,  "Hreflang Missing Return Links"),
        (False, " - Hreflang annotations should be reciprocal, meaning alternate pages must "
                "reference each other. Missing return links can prevent search engines from "
                "confirming the hreflang relationship between alternate pages."),
    ],
    5: [
        (True,  "Issue Summary\n"),
        (False, "In the following report, we covered hreflang_no_index_return_links issue:\n\n"),
        (True,  "Hreflang Noindex Return Links"),
        (False, " \u2013 URLs where the alternate hreflang page is marked as non-indexable using a "
                "noindex directive. These issues can prevent search engines from properly "
                "processing hreflang relationships."),
    ],
    6: [
        (True,  "Issue Summary\n"),
        (False, "In the following report we covered Hreflang multiple entries issue:\n\n"),
        (True,  "Hreflang Multiple Entries"),
        (False, " \u2013 URLs containing multiple hreflang entries for the same language or "
                "regional code. Multiple hreflang entries for the same language can create "
                "conflicting hreflang signals."),
    ],
    7: [
        (True,  "Issue Summary\n"),
        (True,  "Hreflang non canonical return links"),
        (False, " - URLs with hreflang return links pointing to non-canonical URLs. "
                "Non-canonical return links can create conflicting signals for search engines."),
    ],
}

SUMMARY_CELL_RANGES = [
    ("A2",  "A2:F11"),
    ("A2",  "A2:F9"),
    ("A1",  "A1:F4"),
    ("A1",  "A1:F5"),
    ("A1",  "A1:F9"),
    ("A1",  "A1:F10"),
    ("A1",  "A1:F8"),
    ("A3",  "A3:F9"),
]

T2_HDR_GRAY_RANGES = [
    (24, 1, 2, 3, 6),
    (21, 1, 2, 3, 4),
    None,
    None,
    None,
    None,
    None,
    None,
]


def apply_openpyxl_fixes(path, t3_hdr_rows):
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText

    wb  = load_workbook(path)
    bld = InlineFont(b=True,  rFont="Calibri", sz=10, color="FF000000")
    nrm = InlineFont(b=False, rFont="Calibri", sz=10, color="FF000000")
    gray  = _gray_fill()
    white = _white_fill()
    bdr   = _border()

    def apply_row_fill(ws_op, xl_row, col_start, col_end, fill, font_size=9):
        for col in range(col_start, col_end + 1):
            c = ws_op.cell(row=xl_row, column=col)
            c.fill   = fill
            c.border = bdr
            c.font   = Font(name="Calibri", size=font_size, bold=True)

    for sheet_idx, ws_op in enumerate(wb.worksheets):
        # Rich summary
        pairs     = SUMMARY_PAIRS.get(sheet_idx, [])
        cell_ref, merge_range = SUMMARY_CELL_RANGES[sheet_idx]
        if pairs:
            try:
                ws_op.merge_cells(merge_range)
            except Exception:
                pass
            c = ws_op[cell_ref]
            blocks = [TextBlock(bld if bold else nrm, txt) for bold, txt in pairs]
            c.value     = CellRichText(*blocks)
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border    = bdr

        # T2 header fills for sheets 1 and 2
        t2_rng = T2_HDR_GRAY_RANGES[sheet_idx]
        if t2_rng:
            xl_row, g_start, g_end, w_start, w_end = t2_rng
            apply_row_fill(ws_op, xl_row, g_start, g_end, gray,  font_size=8)
            apply_row_fill(ws_op, xl_row, w_start, w_end, white, font_size=8)

        # T3 header fill (dynamic row)
        t3_hdr_0idx = t3_hdr_rows[sheet_idx]
        t3_xl = t3_hdr_0idx + 1
        max_col = 0
        for cell in ws_op[t3_xl]:
            if cell.value is not None:
                max_col = cell.column
        if max_col > 0:
            apply_row_fill(ws_op, t3_xl, 1, max_col, gray, font_size=9)

    wb.save(path)


# ── MAIN ENTRY POINT ──────────────────────────────────────────────────────────
def build_hreflang_masterfile(crawl_id: str, domain: str, report_path: str) -> bytes:
    rulebook = load_rulebook(domain)

    df_int = pd.read_csv(os.path.join(report_path, "internal_all.csv"),
                         encoding="utf-8", low_memory=False) \
             if os.path.exists(os.path.join(report_path, "internal_all.csv")) else pd.DataFrame()

    df_gsc_path = os.path.join(report_path, "search_console_all.csv")
    df_gsc = pd.read_csv(df_gsc_path, encoding="utf-8", low_memory=False) \
             if os.path.exists(df_gsc_path) else pd.DataFrame()

    df_ga_path = os.path.join(report_path, "analytics_all.csv")
    df_ga = pd.read_csv(df_ga_path, encoding="utf-8", low_memory=False) \
            if os.path.exists(df_ga_path) else pd.DataFrame()

    internal_map    = build_internal_map(df_int)
    gsc_map         = build_gsc_map(df_gsc)
    ga_map          = build_ga_map(df_ga)
    df_filt         = filter_indexable_200_html(df_int)
    total_indexable = max(len(df_filt) if not df_filt.empty else 1, 1)

    buf  = io.BytesIO()
    wb   = xlsxwriter.Workbook(buf, {"in_memory": True, "nan_inf_to_errors": True,
                                     "strings_to_urls": False})
    fmts = make_formats(wb)

    args = (fmts, report_path, rulebook, internal_map, gsc_map, ga_map, total_indexable)

    build_sheet1(wb, *args)
    _, _, s2_t3_hdr, _ = build_sheet2(wb, *args)
    s3_t3_hdr, _ = build_sheet3(wb, *args)
    s4_t3_hdr, _ = build_sheet4(wb, *args)
    s5_t3_hdr, _ = build_sheet5(wb, *args)
    s6_t3_hdr, _ = build_sheet6(wb, *args)
    s7_t3_hdr, _ = build_sheet7(wb, *args)
    s8_t3_hdr, _ = build_sheet8(wb, *args)

    wb.close()
    xlsx_bytes = buf.getvalue()

    # Pass 2: openpyxl for rich-text summaries + gray T3 header fills
    t3_hdr_rows = [32, s2_t3_hdr, s3_t3_hdr, s4_t3_hdr,
                   s5_t3_hdr, s6_t3_hdr, s7_t3_hdr, s8_t3_hdr]

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
        tmp.write(xlsx_bytes)

    try:
        apply_openpyxl_fixes(tmp_path, t3_hdr_rows)
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
