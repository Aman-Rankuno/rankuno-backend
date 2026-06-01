import os
import io
import zipfile
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.config import settings
from app.services.rulebook import load_rulebook, classify_url

BLACK_FONT   = Font(name="Arial", size=9, color="FF000000")
HEADER_FONT  = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
HEADER_FILL  = PatternFill("solid", fgColor="FFFF0000")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
NO_FILL      = PatternFill(fill_type=None)
THIN_BORDER  = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2, "N/A": 3}

# ── Table 1 fixed rows (from template inspection) ─────────────────────────────
T1_DATA_ROW = 22   # #Affected URLs
T1_PCT_ROW  = 23   # % Share

# Table 1 columns: A=label, B=Short, C=Long, D=Missing, E=Duplicate,
#                  F=Multiple, G=Outside Head, H=Total Pages
T1_SHORT_COL     = 2
T1_LONG_COL      = 3
T1_MISSING_COL   = 4
T1_DUPLICATE_COL = 5
T1_MULTIPLE_COL  = 6
T1_OUTSIDE_COL   = 7
T1_TOTAL_COL     = 8

# ── Table 2 starts at row 28 ──────────────────────────────────────────────────
T2_DATA_START_ROW = 28
# Table 2 columns: A=Page Theme, B=Total Pages, C=Priority, D=Short,
#                  E=Long, F=Missing, G=Duplicate, H=Multiple, I=Outside Head
T2_THEME_COL     = 1
T2_TOTAL_COL     = 2
T2_PRIORITY_COL  = 3
T2_SHORT_COL     = 4
T2_LONG_COL      = 5
T2_MISSING_COL   = 6
T2_DUPLICATE_COL = 7
T2_MULTIPLE_COL  = 8
T2_OUTSIDE_COL   = 9

# ── Table 3 columns ───────────────────────────────────────────────────────────
T3_COLS = [
    "Address", "Page Theme 1", "Page Theme 2", "Content Type",
    "Status Code", "Indexability", "Meta Description",
    "Meta Description Pixel Width",
    "Short", "Long", "Missing", "Duplicate", "Multiple", "Outside Head",
]

ISSUE_KEYS = ["Short", "Long", "Missing", "Duplicate", "Multiple", "Outside Head"]

ISSUE_COL_MAP = {
    "Short":        T1_SHORT_COL,
    "Long":         T1_LONG_COL,
    "Missing":      T1_MISSING_COL,
    "Duplicate":    T1_DUPLICATE_COL,
    "Multiple":     T1_MULTIPLE_COL,
    "Outside Head": T1_OUTSIDE_COL,
}

T2_ISSUE_COL_MAP = {
    "Short":        T2_SHORT_COL,
    "Long":         T2_LONG_COL,
    "Missing":      T2_MISSING_COL,
    "Duplicate":    T2_DUPLICATE_COL,
    "Multiple":     T2_MULTIPLE_COL,
    "Outside Head": T2_OUTSIDE_COL,
}

KEEP_FROM_TEMPLATE = {
    'xl/drawings/drawing1.xml',
    'xl/drawings/vmlDrawing1.vml',
    'xl/comments1.xml',
    'xl/persons/person.xml',
    'xl/documenttasks/documenttask1.xml',
    'xl/threadedComments/threadedComment1.xml',
}


def _merge_with_drawings(template_path: str, openpyxl_buf: io.BytesIO) -> bytes:
    openpyxl_buf.seek(0)
    with zipfile.ZipFile(template_path, 'r') as t_zip:
        template_sheet1_rels = t_zip.read('xl/worksheets/_rels/sheet1.xml.rels')
        template_sheet2_rels = (
            t_zip.read('xl/worksheets/_rels/sheet2.xml.rels')
            if 'xl/worksheets/_rels/sheet2.xml.rels' in t_zip.namelist()
            else None
        )

    final_buf = io.BytesIO()
    with zipfile.ZipFile(openpyxl_buf, 'r') as o_zip:
        with zipfile.ZipFile(final_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:
            for item in o_zip.namelist():
                if item in ('xl/comments/comment1.xml', 'xl/drawings/commentsDrawing1.vml'):
                    continue
                if item == 'xl/worksheets/_rels/sheet1.xml.rels':
                    out_zip.writestr(item, template_sheet1_rels)
                else:
                    out_zip.writestr(item, o_zip.read(item))
            with zipfile.ZipFile(template_path, 'r') as t_zip:
                for item in KEEP_FROM_TEMPLATE:
                    if item in t_zip.namelist() and item not in o_zip.namelist():
                        out_zip.writestr(item, t_zip.read(item))
                if template_sheet2_rels and 'xl/worksheets/_rels/sheet2.xml.rels' not in o_zip.namelist():
                    out_zip.writestr('xl/worksheets/_rels/sheet2.xml.rels', template_sheet2_rels)

    final_buf.seek(0)
    return final_buf.read()


def build_meta_description_masterfile(crawl_id: str, domain: str, report_path: str) -> bytes:
    template_path = os.path.join(settings.TEMPLATES_DIR, "Meta_Description_Issues.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    rulebook = load_rulebook(domain)

    def load_csv(filename):
        path = os.path.join(report_path, filename)
        if os.path.exists(path):
            try:
                return pd.read_csv(path, encoding="utf-8", low_memory=False)
            except Exception:
                return pd.DataFrame()
        return pd.DataFrame()

    df_internal_all = load_csv("internal_all.csv")
    df_gsc          = load_csv("search_console_all.csv")
    df_ga           = load_csv("analytics_all.csv")

    if df_internal_all.empty:
        raise ValueError("internal_all.csv not found or empty")

    def url_set(filename):
        df = load_csv(filename)
        if df.empty:
            return set()
        col = next((c for c in df.columns if c.lower() == "address"), None)
        return set(df[col].dropna().astype(str)) if col else set()

    short_urls     = url_set("meta_description_below_400_pixels.csv")
    long_urls      = url_set("meta_description_over_985_pixels.csv")
    missing_urls   = url_set("meta_description_missing.csv")
    duplicate_urls = url_set("meta_description_duplicate.csv")
    multiple_urls  = url_set("meta_description_multiple.csv")
    outside_urls   = url_set("meta_description_outside_head.csv")

    gsc_map = {}
    if not df_gsc.empty:
        a   = next((c for c in df_gsc.columns if c.lower() == "address"), None)
        imp = next((c for c in df_gsc.columns if "impression" in c.lower()), None)
        clk = next((c for c in df_gsc.columns if "click" in c.lower()), None)
        if a:
            for _, r in df_gsc.iterrows():
                gsc_map[str(r[a])] = {
                    "impressions": r.get(imp, 0) if imp else 0,
                    "clicks":      r.get(clk, 0) if clk else 0,
                }

    ga_map = {}
    if not df_ga.empty:
        a = next((c for c in df_ga.columns if c.lower() == "address"), None)
        s = next((c for c in df_ga.columns if "session" in c.lower()), None)
        if a and s:
            for _, r in df_ga.iterrows():
                ga_map[str(r[a])] = r.get(s, 0)

    def get_col(df, *names):
        for n in names:
            m = next((c for c in df.columns if c.lower() == n.lower()), None)
            if m:
                return m
        return None

    ia_addr   = get_col(df_internal_all, "Address")
    ia_status = get_col(df_internal_all, "Status Code")
    ia_idx    = get_col(df_internal_all, "Indexability")
    ia_ct     = get_col(df_internal_all, "Content Type")
    ia_meta   = get_col(df_internal_all, "Meta Description 1", "Meta Description")
    ia_pixel  = get_col(df_internal_all, "Meta Description 1 Pixel Width", "Meta Description Pixel Width")

    mask = (
        (df_internal_all[ia_status].astype(str) == "200") &
        (df_internal_all[ia_idx].astype(str).str.lower() == "indexable") &
        (df_internal_all[ia_ct].astype(str).str.lower().str.contains("text/html", na=False))
    )
    total_indexable = int(mask.sum())
    df_base = df_internal_all[mask].copy()

    df_base["_url"] = df_base[ia_addr].astype(str)
    classifications = [classify_url(url, rulebook) for url in df_base["_url"]]
    df_base["Page Theme 1"] = [c[0] for c in classifications]
    df_base["Page Theme 2"] = [c[1] if c[1] else "-" for c in classifications]
    df_base["_priority"]    = [c[3] for c in classifications]

    df_base["Short"]        = df_base["_url"].isin(short_urls).map({True: "Yes", False: "No"})
    df_base["Long"]         = df_base["_url"].isin(long_urls).map({True: "Yes", False: "No"})
    df_base["Missing"]      = df_base["_url"].isin(missing_urls).map({True: "Yes", False: "No"})
    df_base["Duplicate"]    = df_base["_url"].isin(duplicate_urls).map({True: "Yes", False: "No"})
    df_base["Multiple"]     = df_base["_url"].isin(multiple_urls).map({True: "Yes", False: "No"})
    df_base["Outside Head"] = df_base["_url"].isin(outside_urls).map({True: "Yes", False: "No"})

    df_base["Impressions"]      = df_base["_url"].map(lambda u: gsc_map.get(u, {}).get("impressions", 0))
    df_base["Clicks"]           = df_base["_url"].map(lambda u: gsc_map.get(u, {}).get("clicks", 0))
    df_base["Organic Sessions"] = df_base["_url"].map(lambda u: ga_map.get(u, 0))

    rename_map = {}
    if ia_addr:   rename_map[ia_addr]   = "Address"
    if ia_ct:     rename_map[ia_ct]     = "Content Type"
    if ia_status: rename_map[ia_status] = "Status Code"
    if ia_idx:    rename_map[ia_idx]    = "Indexability"
    if ia_meta:   rename_map[ia_meta]   = "Meta Description"
    if ia_pixel:  rename_map[ia_pixel]  = "Meta Description Pixel Width"

    df_table3 = df_base.rename(columns=rename_map)
    for col in T3_COLS:
        if col not in df_table3.columns:
            df_table3[col] = None

    df_table3 = df_table3[T3_COLS + ["_priority", "Impressions", "Clicks", "Organic Sessions"]].copy()
    df_table3 = df_table3.sort_values("Impressions", ascending=False).reset_index(drop=True)

    issue_counts = {
        "Short":        len(short_urls),
        "Long":         len(long_urls),
        "Missing":      len(missing_urls),
        "Duplicate":    len(duplicate_urls),
        "Multiple":     len(multiple_urls),
        "Outside Head": len(outside_urls),
    }

    theme_data = {}
    for _, row in df_table3.iterrows():
        theme    = row["Page Theme 1"] or "Others"
        priority = row.get("_priority", "Low")
        if theme not in theme_data:
            theme_data[theme] = {"total": 0, "priority": priority, **{k: 0 for k in ISSUE_KEYS}}
        else:
            if PRIORITY_ORDER.get(priority, 2) < PRIORITY_ORDER.get(theme_data[theme]["priority"], 2):
                theme_data[theme]["priority"] = priority
        theme_data[theme]["total"] += 1
        for issue in ISSUE_KEYS:
            if row[issue] == "Yes":
                theme_data[theme][issue] += 1

    sorted_themes = sorted(
        theme_data.items(),
        key=lambda x: PRIORITY_ORDER.get(x[1]["priority"], 2)
    )

    # ── Dynamic row positions based on actual theme count ─────────────────────
    num_themes     = len(sorted_themes)
    t2_last_row    = T2_DATA_START_ROW + num_themes - 1
    t3_label_row   = t2_last_row + 2        # one blank row after Table 2
    t3_header_row  = t3_label_row + 1
    t3_data_start  = t3_header_row + 1

    # ── Load template and write ───────────────────────────────────────────────
    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb['Meta Description Issues']

    # Clear everything from Table 2 data start downward
    ws.auto_filter.ref = None
    for r in range(T2_DATA_START_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.font  = BLACK_FONT
            cell.fill  = NO_FILL

    # ── Fill Table 1 ─────────────────────────────────────────────────────────
    for issue, col in ISSUE_COL_MAP.items():
        cnt = issue_counts[issue]
        ws.cell(T1_DATA_ROW, col).value = cnt if cnt > 0 else None
        ws.cell(T1_DATA_ROW, col).font  = BLACK_FONT
    ws.cell(T1_DATA_ROW, T1_TOTAL_COL).value = total_indexable if total_indexable > 0 else None
    ws.cell(T1_DATA_ROW, T1_TOTAL_COL).font  = BLACK_FONT

    for issue, col in ISSUE_COL_MAP.items():
        cnt = issue_counts[issue]
        pct = round(cnt / total_indexable * 100, 2) if total_indexable > 0 and cnt > 0 else None
        ws.cell(T1_PCT_ROW, col).value = pct
        ws.cell(T1_PCT_ROW, col).font  = BLACK_FONT
    ws.cell(T1_PCT_ROW, T1_TOTAL_COL).value = "-"
    ws.cell(T1_PCT_ROW, T1_TOTAL_COL).font  = BLACK_FONT

    # ── Fill Table 2 (dynamic rows) ───────────────────────────────────────────
    for row_offset, (theme, counts) in enumerate(sorted_themes):
        r           = T2_DATA_START_ROW + row_offset
        total_theme = counts["total"]

        # Write values first
        ws.cell(r, T2_THEME_COL).value    = theme
        ws.cell(r, T2_TOTAL_COL).value    = total_theme
        ws.cell(r, T2_PRIORITY_COL).value = counts["priority"]
        for issue, col in T2_ISSUE_COL_MAP.items():
            cnt = counts[issue]
            ws.cell(r, col).value = (
                f"{cnt} ({round(cnt / total_theme * 100)}%)"
                if cnt > 0 and total_theme > 0 else None
            )

        # Apply formatting after writing
        for col in range(1, T2_OUTSIDE_COL + 1):
            cell           = ws.cell(r, col)
            cell.font      = BLACK_FONT
            cell.border    = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.fill      = NO_FILL

    # ── Write Table 2 label (dynamic position) ────────────────────────────────
    lbl            = ws.cell(t3_label_row, 1)
    lbl.value      = "Table 2"
    lbl.font       = Font(name="Arial", bold=True, size=11, color="FF000000")
    lbl.fill       = NO_FILL

    # ── Write Table 3 headers (dynamic position) ─────────────────────────────
    for col_offset, header in enumerate(T3_COLS):
        cell           = ws.cell(t3_header_row, 1 + col_offset)
        cell.value     = header
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = THIN_BORDER

    # ── Set auto-filter on Table 3 header row ─────────────────────────────────
    last_data_row = t3_data_start + len(df_table3) - 1
    ws.auto_filter.ref = f"A{t3_header_row}:N{last_data_row}"

    # ── Fill Table 3 data ─────────────────────────────────────────────────────
    for row_offset, (_, row) in enumerate(df_table3.iterrows()):
        r = t3_data_start + row_offset
        for col_offset, col_name in enumerate(T3_COLS):
            val = row.get(col_name, "")
            if pd.isna(val) or val == "":
                val = None
            cell       = ws.cell(r, 1 + col_offset)
            cell.value = val
            cell.font  = BLACK_FONT
            cell.fill  = NO_FILL

    openpyxl_buf = io.BytesIO()
    wb.save(openpyxl_buf)
    return _merge_with_drawings(template_path, openpyxl_buf)