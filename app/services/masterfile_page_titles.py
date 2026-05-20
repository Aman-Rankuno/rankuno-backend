import os
import io
import zipfile
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.config import settings
from app.services.rulebook import load_rulebook, classify_url

BLACK_FONT = Font(name="Arial", size=9, color="FF000000")
HEADER_FONT = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="FFFF0000")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
NO_FILL = PatternFill(fill_type=None)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Fixed template row positions
T1_DATA_ROW = 13
T1_PCT_ROW = 14
T2_DATA_START_ROW = 19

# Table 1 column positions
T1_SHORT_COL = 2
T1_LONG_COL = 3
T1_MISSING_COL = 4
T1_DUPLICATE_COL = 5
T1_MULTIPLE_COL = 6
T1_SAME_H1_COL = 7
T1_OUTSIDE_COL = 8
T1_TOTAL_COL = 9

# Table 2 columns
T2_THEME_COL = 1
T2_TOTAL_COL = 2
T2_PRIORITY_COL = 3
T2_SHORT_COL = 4
T2_LONG_COL = 5
T2_MISSING_COL = 6
T2_DUPLICATE_COL = 7
T2_MULTIPLE_COL = 8
T2_SAME_H1_COL = 9
T2_OUTSIDE_COL = 10

T3_COLS = [
    "Address", "Page Theme 1", "Page Theme 2", "Content Type",
    "Status Code", "Indexability", "Title", "Title Pixel Width",
    "Short", "Long", "Missing", "Duplicate", "Multiple",
    "Same as H1", "Outside Head", "Impressions", "Clicks", "Organic Sessions"
]

ISSUE_KEYS = ["Short", "Long", "Missing", "Duplicate", "Multiple", "Same as H1", "Outside Head"]
PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}


def _clear_cell(ws, row, col):
    """Clear a cell's value, fill, and reset font to black."""
    cell = ws.cell(row, col)
    cell.value = None
    cell.fill = NO_FILL
    cell.font = BLACK_FONT


def _write_header_cell(ws, row, col, value):
    cell = ws.cell(row, col, value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN


def _merge_with_drawings(template_path: str, openpyxl_bytes: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(openpyxl_bytes), 'r') as oxl:
        new_sheet = oxl.read('xl/worksheets/sheet1.xml')
        try:
            new_styles = oxl.read('xl/styles.xml')
        except KeyError:
            new_styles = None
        try:
            new_shared = oxl.read('xl/sharedStrings.xml')
        except KeyError:
            new_shared = None

    out_buf = io.BytesIO()
    with zipfile.ZipFile(template_path, 'r') as tmpl:
        with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:
            for item in tmpl.namelist():
                if item == 'xl/worksheets/sheet1.xml':
                    out_zip.writestr(item, new_sheet)
                elif item == 'xl/styles.xml' and new_styles:
                    out_zip.writestr(item, new_styles)
                elif item == 'xl/sharedStrings.xml' and new_shared:
                    out_zip.writestr(item, new_shared)
                else:
                    out_zip.writestr(item, tmpl.read(item))
            if new_shared and 'xl/sharedStrings.xml' not in tmpl.namelist():
                out_zip.writestr('xl/sharedStrings.xml', new_shared)

    out_buf.seek(0)
    return out_buf.read()


def build_page_titles_masterfile(crawl_id: str, domain: str, report_path: str) -> bytes:
    template_path = os.path.join(settings.TEMPLATES_DIR, "Page Titles.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    rulebook = load_rulebook(domain)

    def load_csv(filename):
        path = os.path.join(report_path, filename)
        if os.path.exists(path):
            try:
                return pd.read_csv(path, encoding="utf-8", low_memory=False)
            except Exception:
                return pd.DataFrame()
        return pd.DataFrame()

    df_missing = load_csv("page_titles_missing.csv")
    df_duplicate = load_csv("page_titles_duplicate.csv")
    df_over_pixels = load_csv("page_titles_over_561_pixels.csv")
    df_below_pixels = load_csv("page_titles_below_200_pixels.csv")
    df_same_h1 = load_csv("page_titles_same_as_h1.csv")
    df_multiple = load_csv("page_titles_multiple.csv")
    df_outside_head = load_csv("page_titles_outside_head.csv")
    df_internal_all = load_csv("internal_all.csv")
    df_gsc = load_csv("search_console_all.csv")
    df_ga = load_csv("analytics_all.csv")

    def url_set(df):
        if df.empty:
            return set()
        col = next((c for c in df.columns if c.lower() == "address"), None)
        return set(df[col].dropna().astype(str)) if col else set()

    missing_urls = url_set(df_missing)
    duplicate_urls = url_set(df_duplicate)
    long_urls = url_set(df_over_pixels)
    short_urls = url_set(df_below_pixels)
    same_h1_urls = url_set(df_same_h1)
    multiple_urls = url_set(df_multiple)
    outside_head_urls = url_set(df_outside_head)

    gsc_map = {}
    if not df_gsc.empty:
        a = next((c for c in df_gsc.columns if c.lower() == "address"), None)
        imp = next((c for c in df_gsc.columns if "impression" in c.lower()), None)
        clk = next((c for c in df_gsc.columns if "click" in c.lower()), None)
        if a:
            for _, r in df_gsc.iterrows():
                gsc_map[str(r[a])] = {
                    "impressions": r.get(imp, 0) if imp else 0,
                    "clicks": r.get(clk, 0) if clk else 0,
                }

    ga_map = {}
    if not df_ga.empty:
        a = next((c for c in df_ga.columns if c.lower() == "address"), None)
        s = next((c for c in df_ga.columns if "session" in c.lower()), None)
        if a and s:
            for _, r in df_ga.iterrows():
                ga_map[str(r[a])] = r.get(s, 0)

    if df_internal_all.empty:
        raise ValueError("internal_all.csv not found or empty")

    def get_col(df, *names):
        for n in names:
            m = next((c for c in df.columns if c.lower() == n.lower()), None)
            if m:
                return m
        return None

    ia_addr = get_col(df_internal_all, "Address")
    ia_status = get_col(df_internal_all, "Status Code")
    ia_idx = get_col(df_internal_all, "Indexability")
    ia_ct = get_col(df_internal_all, "Content Type")
    ia_title = get_col(df_internal_all, "Title 1", "Title")
    ia_pixel = get_col(df_internal_all, "Title 1 Pixel Width", "Title Pixel Width")

    mask = (
        (df_internal_all[ia_status].astype(str) == "200") &
        (df_internal_all[ia_idx].astype(str).str.lower() == "indexable") &
        (df_internal_all[ia_ct].astype(str).str.lower().str.contains("text/html", na=False))
    )
    total_indexable = int(mask.sum())
    df_base = df_internal_all[mask].copy()

    rows = []
    for _, row in df_base.iterrows():
        url = str(row.get(ia_addr, "")) if ia_addr else ""
        page_theme1, page_theme2, _, _ = classify_url(url, rulebook)
        gsc = gsc_map.get(url, {})
        rows.append({
            "Address": url,
            "Page Theme 1": page_theme1,
            "Page Theme 2": page_theme2 if page_theme2 else "-",
            "Content Type": row.get(ia_ct, "") if ia_ct else "",
            "Status Code": row.get(ia_status, "") if ia_status else "",
            "Indexability": row.get(ia_idx, "") if ia_idx else "",
            "Title": row.get(ia_title, "") if ia_title else "",
            "Title Pixel Width": row.get(ia_pixel, "") if ia_pixel else "",
            "Short": "Yes" if url in short_urls else "No",
            "Long": "Yes" if url in long_urls else "No",
            "Missing": "Yes" if url in missing_urls else "No",
            "Duplicate": "Yes" if url in duplicate_urls else "No",
            "Multiple": "Yes" if url in multiple_urls else "No",
            "Same as H1": "Yes" if url in same_h1_urls else "No",
            "Outside Head": "Yes" if url in outside_head_urls else "No",
            "Impressions": gsc.get("impressions", 0),
            "Clicks": gsc.get("clicks", 0),
            "Organic Sessions": ga_map.get(url, 0),
        })

    df_table3 = pd.DataFrame(rows)
    if not df_table3.empty:
        df_table3 = df_table3.sort_values("Impressions", ascending=False).reset_index(drop=True)

    issue_counts = {
        "Short": len(short_urls), "Long": len(long_urls),
        "Missing": len(missing_urls), "Duplicate": len(duplicate_urls),
        "Multiple": len(multiple_urls), "Same as H1": len(same_h1_urls),
        "Outside Head": len(outside_head_urls),
    }

    theme_data = {}
    for _, row in df_table3.iterrows():
        theme = row["Page Theme 1"] or "Others"
        _, _, _, priority = classify_url(row["Address"], rulebook)
        if theme not in theme_data:
            theme_data[theme] = {
                "total": 0, "priority": priority,
                "Short": 0, "Long": 0, "Missing": 0, "Duplicate": 0,
                "Multiple": 0, "Same as H1": 0, "Outside Head": 0
            }
        theme_data[theme]["total"] += 1
        for issue in ISSUE_KEYS:
            if row[issue] == "Yes":
                theme_data[theme][issue] += 1

    sorted_themes = sorted(
        theme_data.items(),
        key=lambda x: PRIORITY_ORDER.get(x[1]["priority"], 2)
    )

    # Dynamic Table 3 position: T2 data start + num themes + 1 blank row
    num_themes = len(sorted_themes)
    t3_label_row = T2_DATA_START_ROW + num_themes + 1
    t3_header_row = t3_label_row + 1
    t3_data_start = t3_header_row + 1

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Clear everything from T2_DATA_START_ROW to end, including fills
    for r in range(T2_DATA_START_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            _clear_cell(ws, r, c)

    # Reset Table 1 data rows font only (not fill, headers stay)
    for r in range(T1_DATA_ROW, T1_PCT_ROW + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).font = BLACK_FONT

    # Fill Table 1: #Affected URLs
    col_map = {
        "Short": T1_SHORT_COL, "Long": T1_LONG_COL, "Missing": T1_MISSING_COL,
        "Duplicate": T1_DUPLICATE_COL, "Multiple": T1_MULTIPLE_COL,
        "Same as H1": T1_SAME_H1_COL, "Outside Head": T1_OUTSIDE_COL,
    }
    for issue, col in col_map.items():
        cnt = issue_counts[issue]
        c = ws.cell(T1_DATA_ROW, col)
        c.value = cnt if cnt > 0 else None
        c.font = BLACK_FONT

    c = ws.cell(T1_DATA_ROW, T1_TOTAL_COL)
    c.value = total_indexable if total_indexable > 0 else None
    c.font = BLACK_FONT

    # Fill Table 1: % Share
    for issue, col in col_map.items():
        cnt = issue_counts[issue]
        pct = f"{round(cnt / total_indexable * 100, 2)}" if total_indexable > 0 and cnt > 0 else None
        c = ws.cell(T1_PCT_ROW, col)
        c.value = pct
        c.font = BLACK_FONT

    ws.cell(T1_PCT_ROW, T1_TOTAL_COL).value = "-"
    ws.cell(T1_PCT_ROW, T1_TOTAL_COL).font = BLACK_FONT

    # Fill Table 2
    t2_issue_cols = {
        "Short": T2_SHORT_COL, "Long": T2_LONG_COL, "Missing": T2_MISSING_COL,
        "Duplicate": T2_DUPLICATE_COL, "Multiple": T2_MULTIPLE_COL,
        "Same as H1": T2_SAME_H1_COL, "Outside Head": T2_OUTSIDE_COL,
    }
    for row_offset, (theme, counts) in enumerate(sorted_themes):
        r = T2_DATA_START_ROW + row_offset
        total_theme = counts["total"]
        for col in range(1, 11):
            ws.cell(r, col).border = THIN_BORDER
            ws.cell(r, col).alignment = CENTER_ALIGN
            ws.cell(r, col).fill = NO_FILL
        ws.cell(r, T2_THEME_COL).value = theme
        ws.cell(r, T2_THEME_COL).font = BLACK_FONT
        ws.cell(r, T2_TOTAL_COL).value = total_theme
        ws.cell(r, T2_TOTAL_COL).font = BLACK_FONT
        ws.cell(r, T2_PRIORITY_COL).value = counts["priority"]
        ws.cell(r, T2_PRIORITY_COL).font = BLACK_FONT
        for issue, col in t2_issue_cols.items():
            cnt = counts[issue]
            val = f"{cnt} ({round(cnt / total_theme * 100)}%)" if cnt > 0 and total_theme > 0 else None
            ws.cell(r, col).value = val
            ws.cell(r, col).font = BLACK_FONT

    # Write Table 3 label
    c = ws.cell(t3_label_row, 1, "Table 3")
    c.font = Font(name="Arial", bold=True, size=11, color="FF000000")
    c.fill = NO_FILL

    # Write Table 3 headers with red fill
    for col_offset, header in enumerate(T3_COLS):
        _write_header_cell(ws, t3_header_row, 1 + col_offset, header)

    # Fill Table 3 data
    for row_offset, (_, row) in enumerate(df_table3.iterrows()):
        r = t3_data_start + row_offset
        for col_offset, col_name in enumerate(T3_COLS):
            val = row.get(col_name, "")
            if pd.isna(val) or val == "":
                val = None
            c = ws.cell(r, 1 + col_offset)
            c.value = val
            c.font = BLACK_FONT
            c.fill = NO_FILL

    step1_buf = io.BytesIO()
    wb.save(step1_buf)

    return _merge_with_drawings(template_path, step1_buf.getvalue())