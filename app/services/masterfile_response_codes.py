import os
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.rulebook import load_rulebook, classify_url

RED_FILL = PatternFill("solid", fgColor="FFFF0000")
NORMAL_FONT = Font(name="Arial", size=9)
BOLD_FONT = Font(name="Arial", bold=True, size=9)
BOLD_FONT_11 = Font(name="Arial", bold=True, size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}


def _red_header(cell, value):
    cell.value = value
    cell.fill = RED_FILL
    cell.font = BOLD_FONT
    cell.alignment = CENTER
    cell.border = THIN


def _plain(cell, value, bold=False):
    cell.value = value
    cell.font = BOLD_FONT if bold else NORMAL_FONT
    cell.alignment = LEFT
    cell.border = THIN


def build_response_codes_masterfile(crawl_id: str, domain: str, report_path: str) -> bytes:
    rulebook = load_rulebook(domain)

    def load_csv(filename):
        path = os.path.join(report_path, filename)
        if os.path.exists(path):
            try:
                return pd.read_csv(path, encoding="utf-8", low_memory=False)
            except Exception:
                return pd.DataFrame()
        return pd.DataFrame()

    df_3xx = load_csv("response_codes_internal_redirection_(3xx).csv")
    df_4xx = load_csv("response_codes_internal_client_error_(4xx).csv")
    df_5xx = load_csv("response_codes_internal_server_error_(5xx).csv")
    df_no_response = load_csv("response_codes_internal_no_response.csv")
    df_redirect_chain = load_csv("response_codes_internal_redirect_chain.csv")
    df_redirect_loop = load_csv("response_codes_internal_redirect_loop.csv")
    df_blocked = load_csv("response_codes_internal_blocked_by_robots_txt.csv")
    df_internal_all = load_csv("internal_all.csv")
    df_gsc = load_csv("search_console_all.csv")
    df_ga = load_csv("analytics_all.csv")

    # Total URLs from internal_all (all content types including CSS, PDF, images)
    total_crawled = len(df_internal_all) if not df_internal_all.empty else 0

    # Tag error types
    def tag(df, label):
        if df.empty:
            return df
        df = df.copy()
        df["_error_type"] = label
        return df

    df_3xx_t = tag(df_3xx, "3xx")
    df_4xx_t = tag(df_4xx, "4xx")
    df_5xx_t = tag(df_5xx, "5xx")
    df_no_t = tag(df_no_response, "No response code")
    df_blocked_t = tag(df_blocked, "Blocked by robots")

    # Redirect chain/loop lookup - only for 3xx addresses
    chain_urls = set()
    loop_urls = set()
    if not df_redirect_chain.empty:
        col = next((c for c in df_redirect_chain.columns if c.lower() == "address"), None)
        if col:
            chain_urls = set(df_redirect_chain[col].dropna().astype(str))
    if not df_redirect_loop.empty:
        col = next((c for c in df_redirect_loop.columns if c.lower() == "address"), None)
        if col:
            loop_urls = set(df_redirect_loop[col].dropna().astype(str))

    # GSC lookup
    gsc_map = {}
    if not df_gsc.empty:
        a = next((c for c in df_gsc.columns if c.lower() == "address"), None)
        imp = next((c for c in df_gsc.columns if "impression" in c.lower()), None)
        clk = next((c for c in df_gsc.columns if "click" in c.lower()), None)
        if a:
            for _, r in df_gsc.iterrows():
                gsc_map[str(r[a])] = {
                    "impressions": r.get(imp, 0) if imp else 0,
                    "clicks": r.get(clk, 0) if clk else 0,
                }

    # GA lookup
    ga_map = {}
    if not df_ga.empty:
        a = next((c for c in df_ga.columns if c.lower() == "address"), None)
        s = next((c for c in df_ga.columns if "session" in c.lower()), None)
        if a and s:
            for _, r in df_ga.iterrows():
                ga_map[str(r[a])] = r.get(s, 0)

    # Combine frames
    frames = [df for df in [df_3xx_t, df_4xx_t, df_5xx_t, df_no_t, df_blocked_t] if not df.empty]
    if not frames:
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "No response code data found for this crawl."
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    df_combined = pd.concat(frames, ignore_index=True)

    def get_col(df, *names):
        for n in names:
            m = next((c for c in df.columns if c.lower() == n.lower()), None)
            if m:
                return m
        return None

    addr_col = get_col(df_combined, "Address")
    sc_col = get_col(df_combined, "Status Code")
    st_col = get_col(df_combined, "Status")
    idx_col = get_col(df_combined, "Indexability")
    idx_st_col = get_col(df_combined, "Indexability Status")
    ct_col = get_col(df_combined, "Content Type")
    inl_col = get_col(df_combined, "Inlinks")
    ru_col = get_col(df_combined, "Redirect URL")
    rt_col = get_col(df_combined, "Redirect Type")

    # Build Table 3 rows
    rows = []
    for _, row in df_combined.iterrows():
        url = str(row.get(addr_col, "")) if addr_col else ""
        error_type = str(row.get("_error_type", ""))
        page_theme1, page_theme2 = classify_url(url, rulebook)
        gsc = gsc_map.get(url, {})

        # Redirect chain/loop only for 3xx
        if error_type == "3xx":
            redirect_chain = "TRUE" if url in chain_urls else ""
            redirect_loop = "TRUE" if url in loop_urls else ""
        else:
            redirect_chain = ""
            redirect_loop = ""

        rows.append({
            "Error type": error_type,
            "Address": url,
            "Page Theme 1": page_theme1,
            "Page Theme 2": page_theme2,
            "Content Type": row.get(ct_col, "") if ct_col else "",
            "Status Code": row.get(sc_col, "") if sc_col else "",
            "Status": row.get(st_col, "") if st_col else "",
            "Indexability": row.get(idx_col, "") if idx_col else "",
            "Indexability Status": row.get(idx_st_col, "") if idx_st_col else "",
            "Inlinks": row.get(inl_col, 0) if inl_col else 0,
            "Redirect URL": row.get(ru_col, "") if ru_col else "",
            "Redirect Type": row.get(rt_col, "") if rt_col else "",
            "Redirect chain": redirect_chain,
            "Redirect loop": redirect_loop,
            "Impressions": gsc.get("impressions", 0),
            "Clicks": gsc.get("clicks", 0),
            "Organic Sessions": ga_map.get(url, 0),
        })

    df_table3 = pd.DataFrame(rows)

    # Sort Table 3 by Impressions descending
    df_table3 = df_table3.sort_values("Impressions", ascending=False).reset_index(drop=True)

    # Table 1 summary
    sc_keys = ["301", "302", "307", "308", "400", "401", "403", "404",
               "500", "502", "503", "504", "No response code"]
    priority_map = {
        "301": "Medium", "302": "High", "307": "High", "308": "Medium",
        "400": "High", "401": "High", "403": "High", "404": "High",
        "500": "High", "502": "High", "503": "High", "504": "High",
        "No response code": "High",
    }
    sc_counts = {k: 0 for k in sc_keys}
    for _, row in df_table3.iterrows():
        sc = str(row["Status Code"])
        et = str(row["Error type"])
        if et == "No response code":
            sc_counts["No response code"] += 1
        elif sc in sc_counts:
            sc_counts[sc] += 1

    # Table 2 - page theme wise, sorted by priority
    theme_counts = {}
    for _, row in df_table3.iterrows():
        theme = row["Page Theme 1"] or "Others"
        sc = str(row["Status Code"])
        et = str(row["Error type"])
        if theme not in theme_counts:
            theme_counts[theme] = {"total": 0, "_priority": "Low"}
            for k in sc_keys:
                theme_counts[theme][k] = 0
        theme_counts[theme]["total"] += 1
        if et == "No response code":
            theme_counts[theme]["No response code"] += 1
        elif sc in theme_counts[theme]:
            theme_counts[theme][sc] += 1

    # Sort themes by priority (High first)
    sorted_themes = sorted(
        theme_counts.items(),
        key=lambda x: PRIORITY_ORDER.get(x[1].get("_priority", "Low"), 2)
    )

    # Table 4
    chain_count = len(df_redirect_chain) if not df_redirect_chain.empty else 0
    loop_count = len(df_redirect_loop) if not df_redirect_loop.empty else 0
    total_3xx = len(df_3xx) if not df_3xx.empty else 0

    # Note text
    note_text = (
        'Note regarding 3xx redirection chain and loop -\n\n'
        '1. The "Source" column is common in both internal_redirection_inlinks(3XX) & redirect_chain reports '
        'and represents the page where the internal link was found.\n'
        '2. In the Redirect Chains report, the "Address" column represents the initially requested redirected URL '
        '(equivalent to the destination URL in Internal Redirection Inlinks reports).\n'
        '3. The "Final Address" column is available in the Redirect Chains report and represents the final resolved '
        'destination URL after all redirect hops.\n'
        '4. The Internal Redirection Inlinks (3xx) report contains only: Source, Destination, Anchor Text, Alt Text, '
        'Link Position, Link Path, and other link-level details but does not contain the final resolved redirect URL.\n'
        '5. To identify the final destination URL for redirected URLs from the Internal Redirection Inlinks (3xx) '
        'report, it is recommended to refer to: Internal_All OR Redirect Chains report.\n'
        '6. The Internal_All file provides an easier way to trace redirected URLs and validate final response codes, '
        'indexability, and destination URLs across the crawl dataset.'
    )

    # Legend data
    legend_rows = [
        ("Column", "Purpose", "Notes"),
        ("Error type", "Issue category", "3xx / 4xx / 5xx / No Response"),
        ("Address", "Impacted URL", "Main issue URL"),
        ("Page Theme 1", "URL template/page type", "Source classification"),
        ("Page Theme 2", "Locale/language/category", "Slicer base"),
        ("Content Type", "MIME type", "Useful for assets"),
        ("Status Code", "HTTP response", "Technical issue"),
        ("Status", "Response meaning", "Not Found / Redirected"),
        ("Indexability", "Crawl/index state", "Indexable/non-indexable"),
        ("Indexability Status", "Why non-indexable", "Redirected/Client Error"),
        ("Inlinks", "Internal links count", "Priority indicator"),
        ("Redirect URL", "Final redirect URL", "Only for 3xx"),
        ("Redirect Type", "Redirect classification", "HTTP Redirect"),
        ("Redirect chain", "TRUE/FALSE", "From redirect chain lookup"),
        ("Redirect loop", "TRUE/FALSE", "From redirect loop lookup"),
        ("Impressions", "GSC visibility", "SEO impact"),
        ("Clicks", "GSC traffic", "Business impact"),
        ("Organic Sessions", "Analytics impact", "User impact"),
    ]

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Response Codes - Internal"

    # ── Slicer label rows (rows 7 and 12 in template, we use 1 and 2)
    slicer_fill = PatternFill("solid", fgColor="FFFF0000")
    for r in [1, 2]:
        c = ws.cell(r, 1, "Slicer for below mentioned both table(Config should be based on page theme 2)")
        c.fill = slicer_fill
        c.font = BOLD_FONT
        c.alignment = LEFT

    # ── Note in column U row 1
    note_cell = ws.cell(1, 21, note_text)
    note_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    note_cell.font = NORMAL_FONT
    ws.row_dimensions[1].height = 80

    # ── TABLE 1 starts at row 4, TABLE 4 beside it at column 17 (Q)
    T1_ROW = 4
    T4_COL = 17  # column Q

    ws.cell(T1_ROW, 1, "Table 1").font = BOLD_FONT_11
    ws.cell(T1_ROW, T4_COL, "Table 4").font = BOLD_FONT_11

    # Table 1 header row (row 5)
    H_ROW = T1_ROW + 1
    _red_header(ws.cell(H_ROW, 2), "URL Issue Types")
    for i, sc in enumerate(sc_keys):
        ws.cell(H_ROW, 3 + i, sc).font = BOLD_FONT
        ws.cell(H_ROW, 3 + i).border = THIN
        ws.cell(H_ROW, 3 + i).alignment = CENTER

    # Table 4 header row
    _red_header(ws.cell(H_ROW, T4_COL), "URL Issue Types")
    ws.cell(H_ROW, T4_COL + 1, "Redirect chain").font = BOLD_FONT
    ws.cell(H_ROW, T4_COL + 1).border = THIN
    ws.cell(H_ROW, T4_COL + 2, "Redirect loop").font = BOLD_FONT
    ws.cell(H_ROW, T4_COL + 2).border = THIN

    # Table 1 rows
    row_labels = [
        ("Issue Priority", {sc: priority_map.get(sc, "High") for sc in sc_keys}),
        ("#Affected URLs", {sc: sc_counts[sc] if sc_counts[sc] > 0 else "" for sc in sc_keys}),
        ("% Share against Total URLs Crawled", {
            sc: f"{round(sc_counts[sc] / total_crawled * 100, 1)}%" if total_crawled > 0 and sc_counts[sc] > 0 else ""
            for sc in sc_keys
        }),
    ]

    # Table 4 rows
    t4_values = [
        ("Issue Priority", "High", "High"),
        ("#Affected URLs", chain_count if chain_count > 0 else "", loop_count if loop_count > 0 else ""),
        ("% share against Total 3xx URLs Crawled",
         f"{round(chain_count / total_3xx * 100, 1)}%" if total_3xx > 0 and chain_count > 0 else "",
         f"{round(loop_count / total_3xx * 100, 1)}%" if total_3xx > 0 and loop_count > 0 else ""),
    ]

    for i, (label, values) in enumerate(row_labels):
        r = H_ROW + 1 + i
        _red_header(ws.cell(r, 2), label)
        for j, sc in enumerate(sc_keys):
            _plain(ws.cell(r, 3 + j), values[sc])
            ws.cell(r, 3 + j).alignment = CENTER

        # Table 4 same row
        _red_header(ws.cell(r, T4_COL), t4_values[i][0])
        _plain(ws.cell(r, T4_COL + 1), t4_values[i][1])
        ws.cell(r, T4_COL + 1).alignment = CENTER
        _plain(ws.cell(r, T4_COL + 2), t4_values[i][2])
        ws.cell(r, T4_COL + 2).alignment = CENTER

    # ── TABLE 2
    T2_ROW = H_ROW + len(row_labels) + 2
    ws.cell(T2_ROW, 1, "Table 2").font = BOLD_FONT_11
    T2_ROW += 1

    t2_title = ws.cell(T2_ROW, 1, "Page Theme Wise URL Analysis")
    t2_title.fill = RED_FILL
    t2_title.font = BOLD_FONT
    t2_title.alignment = LEFT
    T2_ROW += 1

    # Table 2 header
    ws.cell(T2_ROW, 1, "Page Theme 1").font = BOLD_FONT
    ws.cell(T2_ROW, 1).border = THIN
    ws.cell(T2_ROW, 2, "Priority Basis Page Theme 1").font = BOLD_FONT
    ws.cell(T2_ROW, 2).border = THIN
    for i, sc in enumerate(sc_keys):
        ws.cell(T2_ROW, 3 + i, sc).font = BOLD_FONT
        ws.cell(T2_ROW, 3 + i).border = THIN
        ws.cell(T2_ROW, 3 + i).alignment = CENTER
    T2_ROW += 1

    for theme, counts in sorted_themes:
        total_theme = counts["total"]
        _plain(ws.cell(T2_ROW, 1), theme, bold=True)
        _plain(ws.cell(T2_ROW, 2), counts.get("_priority", "High"))
        for i, sc in enumerate(sc_keys):
            cnt = counts.get(sc, 0)
            if cnt > 0 and total_theme > 0:
                pct = round(cnt / total_theme * 100)
                val = f"{cnt} ({pct}%)"
            else:
                val = ""
            _plain(ws.cell(T2_ROW, 3 + i), val)
            ws.cell(T2_ROW, 3 + i).alignment = CENTER
        T2_ROW += 1

    # ── TABLE 3
    T3_ROW = T2_ROW + 2
    ws.cell(T3_ROW, 1, "Table 3").font = BOLD_FONT_11
    T3_ROW += 1

    t3_cols = [
        "Error type", "Address", "Page Theme 1", "Page Theme 2",
        "Content Type", "Status Code", "Status", "Indexability",
        "Indexability Status", "Inlinks", "Redirect URL", "Redirect Type",
        "Redirect chain", "Redirect loop", "Impressions", "Clicks", "Organic Sessions"
    ]
    for i, col in enumerate(t3_cols):
        c = ws.cell(T3_ROW, 1 + i, col)
        c.font = BOLD_FONT
        c.border = THIN
        c.alignment = CENTER

    # Legend header at column V (22) on same row as Table 3 header
    legend_header = ws.cell(T3_ROW, 22, "What does it mean")
    legend_header.fill = PatternFill("solid", fgColor="FFFFFF00")
    legend_header.font = BOLD_FONT
    legend_header.alignment = CENTER
    T3_ROW += 1

    for idx, row in df_table3.iterrows():
        for i, col in enumerate(t3_cols):
            val = row.get(col, "")
            if pd.isna(val):
                val = ""
            _plain(ws.cell(T3_ROW, 1 + i), val)

        # Legend rows alongside data
        if idx < len(legend_rows):
            leg = legend_rows[idx]
            for li, lv in enumerate(leg):
                c = ws.cell(T3_ROW, 22 + li, lv)
                c.font = NORMAL_FONT
                c.border = THIN
                c.alignment = LEFT
        T3_ROW += 1

    # ── Column widths
    col_widths = {
        "A": 18, "B": 65, "C": 22, "D": 22, "E": 30,
        "F": 12, "G": 25, "H": 15, "I": 22, "J": 10,
        "K": 60, "L": 20, "M": 15, "N": 15, "O": 15,
        "P": 12, "Q": 18, "R": 18, "S": 18,
        "V": 22, "W": 30, "X": 35,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()