"""
masterfile_structured_data.py
Structured Data Masterfile - 1 sheet, xlsxwriter + openpyxl two-pass.
Issues: Missing Structured Data, Validation Errors, Validation Warnings, Parse Errors
"""
import os, io, math, tempfile
import pandas as pd
import xlsxwriter
import xlsxwriter.utility
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.services.rulebook import load_rulebook, classify_url as _classify_url

RED   = "#FF0000"
WHITE = "#FFFFFF"
BLACK = "#000000"
GRAY  = "#D9D9D9"
PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2, "N/A": 3}

ISSUES = [
    ("structured_data_missing",             "Missing Structured Data",      "Opportunity", "High",   "structured_data_missing.csv"),
    ("structured_data_validation_errors",   "Validation Errors",            "Issue",       "High",   "structured_data_validation_errors.csv"),
    ("structured_data_validation_warnings", "Validation Warning",           "Warning",     "Low",    "structured_data_validation_warnings.csv"),
    ("structured_data_parse_errors",        "Parse Errors",                 "Issue",       "High",   "structured_data_parse_errors.csv"),
]

SUMMARY_PAIRS = [
    (False, "Issues covered in the following reports are "),
    (True,  "Missing Structured Data, Validation Errors, Validation Warnings, Parse Errors"),
    (False, ".\n\n"),
    (True,  "1. Missing Structured Data"),
    (False, " -  indicates pages without schema implementation and potential opportunities for enhanced search result visibility.\n"),
    (True,  "2. Validation Errors"),
    (False, " -  indicate invalid or unsupported Schema.org properties/types.\n"),
    (True,  "3. Parse Errors"),
    (False, " - indicate malformed structured data syntax that search engines may fail to process correctly.\n"),
    (True,  "4. Validation Warnings"),
    (False, " - highlight deprecated or recommended schema improvements."),
]


def safe_num(v):
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else f
    except Exception:
        return None


def _clean(v):
    s = str(v).strip() if v is not None else ""
    return "-" if (not s or s in ("nan", "None", "NaN")) else s


def _gc(df, *names):
    for n in names:
        m = next((c for c in df.columns if c.lower() == n.lower()), None)
        if m:
            return m
    return None


def load_csv(folder, filename):
    p = os.path.join(folder, filename)
    if os.path.exists(p) and os.path.getsize(p) > 10:
        try:
            return pd.read_csv(p, encoding="utf-8", low_memory=False)
        except Exception:
            try:
                return pd.read_csv(p, encoding="latin-1", low_memory=False)
            except Exception:
                pass
    return pd.DataFrame()

def classify(url, rulebook):
    if not url or url == "-":
        return ("-", "-", "N/A")
    try:
        t1, t2, _, pri = _classify_url(url, rulebook)
        return (t1 or "-", t2 or "-", pri or "N/A")
    except Exception:
        return ("-", "-", "N/A")


def sorted_themes_from_rows(rows, issue_keys):
    theme_data = {}
    for r in rows:
        th  = r.get("Page Theme 1", "-") or "-"
        ik  = r.get("_issue_key", "")
        pri = r.get("_priority", "N/A")
        if th not in theme_data:
            theme_data[th] = {k: 0 for k in issue_keys}
            theme_data[th]["_priority"] = "N/A"
        if ik in issue_keys:
            theme_data[th][ik] += 1
        if PRIORITY_ORDER.get(pri, 3) < PRIORITY_ORDER.get(theme_data[th]["_priority"], 3):
            theme_data[th]["_priority"] = pri
    return sorted(theme_data.items(), key=lambda x: PRIORITY_ORDER.get(x[1]["_priority"], 3))


def make_formats(wb):
    def f(**kw): return wb.add_format(kw)
    return {
        "red_lft":  f(bold=True, font_name="Calibri", font_size=8,  font_color=WHITE, bg_color=RED,   border=1, align="left",   valign="vcenter", text_wrap=True),
        "cell":     f(           font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter"),
        "cell_lft": f(           font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="left",   valign="vcenter"),
        "t1_hdr":   f(bold=True, font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter", text_wrap=True),
        "t3_hdr":   f(bold=True, font_name="Calibri", font_size=9,  font_color=BLACK, bg_color=GRAY,  border=1, align="center", valign="vcenter", text_wrap=True),
        "t3_hdr_l": f(bold=True, font_name="Calibri", font_size=9,  font_color=BLACK, bg_color=GRAY,  border=1, align="left",   valign="vcenter", text_wrap=True),
        "t3_cell":  f(           font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter"),
        "t3_cell_l":f(           font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="left",   valign="vcenter"),
        "num":      f(           font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter", num_format="#,##0"),
        "pct":      f(           font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter", num_format="0.00%"),
        "lbl":      f(bold=True, font_name="Calibri", font_size=11, font_color=BLACK),
        "t2_title": f(bold=True, font_name="Calibri", font_size=8,  font_color=WHITE, bg_color=RED,   border=1, align="center", valign="vcenter"),
        "t2_hdr_l": f(bold=True, font_name="Calibri", font_size=8,  font_color=BLACK, bg_color=GRAY,  border=1, align="left",   valign="vcenter", text_wrap=True),
        "t2_hdr_c": f(bold=True, font_name="Calibri", font_size=8,  font_color=BLACK, bg_color=GRAY,  border=1, align="center", valign="vcenter", text_wrap=True),
        "t2_hdr_w": f(bold=True, font_name="Calibri", font_size=8,  font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter", text_wrap=True),
        "t2_cell":  f(           font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="center", valign="vcenter"),
        "t2_cell_l":f(           font_name="Calibri", font_size=11, font_color=BLACK, bg_color=WHITE, border=1, align="left",   valign="vcenter"),
        "summary":  f(           font_name="Calibri", font_size=10, font_color=BLACK, text_wrap=True, valign="top", border=1),
    }


def build_structured_data_masterfile(crawl_id: str, domain: str, report_path: str) -> bytes:
    rulebook = load_rulebook(domain)

    # Load internal_all for total indexable count
    df_int = load_csv(report_path, "internal_all.csv")
    total_indexable = 1
    if not df_int.empty:
        sc  = _gc(df_int, "Status Code")
        idx = _gc(df_int, "Indexability")
        ct  = _gc(df_int, "Content Type")
        if sc and idx and ct:
            mask = (
                (df_int[sc].astype(str) == "200") &
                (df_int[idx].astype(str).str.lower() == "indexable") &
                (df_int[ct].astype(str).str.contains("text/html", case=False, na=False))
            )
            total_indexable = max(1, int(mask.sum()))

    # Load GSC
    df_gsc = load_csv(report_path, "search_console_all.csv")
    gsc_map = {}
    if not df_gsc.empty:
        a = _gc(df_gsc, "Address")
        imp = next((c for c in df_gsc.columns if "impression" in c.lower()), None)
        clk = next((c for c in df_gsc.columns if "click" in c.lower()), None)
        if a:
            for _, r in df_gsc.iterrows():
                gsc_map[_clean(r[a])] = {
                    "impressions": safe_num(r.get(imp)) if imp else None,
                    "clicks":      safe_num(r.get(clk)) if clk else None,
                }

    # Load GA4
    df_ga = load_csv(report_path, "analytics_all.csv")
    ga_map = {}
    if not df_ga.empty:
        a = _gc(df_ga, "Address")
        s = next((c for c in df_ga.columns if "session" in c.lower()), None)
        if a and s:
            for _, r in df_ga.iterrows():
                ga_map[_clean(r[a])] = safe_num(r.get(s))

    # Build T3 rows for all issues
    issue_keys = [ik for ik, _, _, _, _ in ISSUES]
    all_rows = []

    for ik, label, sev, pri, csv_file in ISSUES:
        df = load_csv(report_path, csv_file)
        if df.empty:
            all_rows.append({
                "_issue_key": ik, "Error Type": label,
                "Address": "-", "Page Theme 1": "-", "Page Theme 2": "-",
                "Errors": None, "Warnings": None, "Total Types": None, "Unique Types": None,
                "Indexability": "-", "Indexability Status": "-",
                "Type-1": "-", "Type-2": "-", "Type-3": "-",
                "Impressions": None, "Clicks": None, "Organic Sessions": None,
                "_priority": "N/A",
            })
            continue

        addr = _gc(df, "Address")
        err  = _gc(df, "Errors")
        warn = _gc(df, "Warnings")
        ttypes = _gc(df, "Total Types")
        utypes = _gc(df, "Unique Types")
        idx_c  = _gc(df, "Indexability")
        idxs_c = _gc(df, "Indexability Status")
        t1_c   = _gc(df, "Type-1")
        # Type-2 and Type-3 only in parse errors
        t2_c   = _gc(df, "Type-2") if ik == "structured_data_parse_errors" else None
        t3_c   = _gc(df, "Type-3") if ik == "structured_data_parse_errors" else None

        if not addr:
            continue

        for _, row in df.iterrows():
            url = _clean(row.get(addr, ""))
            if not url or url == "-":
                continue
            t1, t2, upri = classify(url, rulebook)
            gsc = gsc_map.get(url, {})
            all_rows.append({
                "_issue_key":  ik,
                "Error Type":  label,
                "Address":     url,
                "Page Theme 1": t1,
                "Page Theme 2": t2,
                "Errors":        safe_num(row.get(err))    if err    else None,
                "Warnings":      safe_num(row.get(warn))   if warn   else None,
                "Total Types":   safe_num(row.get(ttypes)) if ttypes else None,
                "Unique Types":  safe_num(row.get(utypes)) if utypes else None,
                "Indexability":        _clean(row.get(idx_c,  "")) if idx_c  else "-",
                "Indexability Status": _clean(row.get(idxs_c, "")) if idxs_c else "-",
                "Type-1": _clean(row.get(t1_c, "")) if t1_c else "-",
                "Type-2": _clean(row.get(t2_c, "")) if t2_c else "-",
                "Type-3": _clean(row.get(t3_c, "")) if t3_c else "-",
                "Impressions":      gsc.get("impressions"),
                "Clicks":           gsc.get("clicks"),
                "Organic Sessions": ga_map.get(url),
                "_priority":        upri,
            })

    # Sort T3 by Impressions descending
    all_rows.sort(key=lambda x: (x["Impressions"] is None, -(x["Impressions"] or 0)))

    # Build themes
    themes = sorted_themes_from_rows(all_rows, issue_keys)

    # Dynamic row positions
    T1_LBL  = 12   # 0-indexed row 12 = Excel row 13
    T1_HDR  = 13
    T1_PRI  = 14
    T1_AFF  = 15
    T1_PCT  = 16
    T2_LBL  = 18
    T2_TTL  = 19
    T2_HDR  = 20
    T2_DATA = 21
    T2_END  = T2_DATA + len(themes) - 1
    T3_LBL  = T2_END + 2
    T3_HDR  = T3_LBL + 1
    T3_DATA = T3_HDR + 1
    T3_XL   = T3_DATA + 1  # Excel 1-based

    # Page Theme 1 is col C (index 2) in T3
    PT1_COL = xlsxwriter.utility.xl_col_to_name(2)  # C

    T3_HDRS = [
        "Error Type", "Address", "Page Theme 1", "Page Theme 2",
        "Errors", "Warnings", "Total Types", "Unique Types",
        "Indexability", "Indexability Status",
        "Impressions", "Clicks", "Organic Sessions",
        "Type-1", "Type-2", "Type-3",
    ]
    # Left-align URL and text cols
    LFT = {0, 1, 8, 9, 13, 14, 15}

    # Build workbook
    buf  = io.BytesIO()
    wb   = xlsxwriter.Workbook(buf, {"in_memory": True, "nan_inf_to_errors": True, "strings_to_urls": False})
    fmts = make_formats(wb)
    ws   = wb.add_worksheet("Structured Data")

    # Column widths
    ws.set_column("A:A", 30)
    ws.set_column("B:B", 50)
    ws.set_column("C:D", 18)
    ws.set_column("E:H", 14)
    ws.set_column("I:J", 18)
    ws.set_column("K:M", 14)
    ws.set_column("N:P", 16)

    # Summary placeholder (rows 1-11, 0-indexed 0-10)
    ws.merge_range(0, 0, 10, 7, "", fmts["summary"])

    # TABLE 1
    ws.write(T1_LBL, 0, "Table 1", fmts["lbl"])
    ws.set_row(T1_HDR, 29)
    ws.write(T1_HDR, 0, "URL Issue Types", fmts["red_lft"])
    for i, (ik, label, sev, pri, csv_file) in enumerate(ISSUES):
        ws.write(T1_HDR, i + 1, label, fmts["t1_hdr"])
    ws.write(T1_PRI, 0, "Issue Priority", fmts["red_lft"])
    for i, (ik, label, sev, pri, csv_file) in enumerate(ISSUES):
        ws.write(T1_PRI, i + 1, pri, fmts["cell"])
    ws.write(T1_AFF, 0, "#Affected URLs", fmts["red_lft"])
    for i, (ik, label, sev, pri, csv_file) in enumerate(ISSUES):
        ws.write_formula(T1_AFF, i + 1,
            '=COUNTIF(A%d:A1048576,"%s")' % (T3_XL, label), fmts["num"])
    ws.set_row(T1_PCT, 21)
    ws.write(T1_PCT, 0, "% Share against Total  HTML URLs Crawled", fmts["red_lft"])
    for i in range(len(ISSUES)):
        cl = xlsxwriter.utility.xl_col_to_name(i + 1)
        ws.write_formula(T1_PCT, i + 1,
            "=%s%d/%d" % (cl, T1_AFF + 1, total_indexable), fmts["pct"])

    # TABLE 2
    ws.write(T2_LBL, 0, "Table 2", fmts["lbl"])
    ws.merge_range(T2_TTL, 0, T2_TTL, len(ISSUES) + 1,
                   "Page Theme Wise URL Analysis ", fmts["t2_title"])
    ws.set_row(T2_HDR, 29)
    ws.write(T2_HDR, 0, "Page Theme 1",               fmts["t2_hdr_l"])
    ws.write(T2_HDR, 1, "Priority Basis Page Theme 1", fmts["t2_hdr_c"])
    for i, (ik, label, sev, pri, csv_file) in enumerate(ISSUES):
        ws.write(T2_HDR, i + 2, label, fmts["t2_hdr_w"])
    for ri, (theme, cnts) in enumerate(themes):
        r = T2_DATA + ri
        ws.set_row(r, 14.5)
        ws.write(r, 0, theme, fmts["t2_cell_l"])
        ws.write(r, 1, cnts["_priority"], fmts["t2_cell"])
        for i, (ik, label, sev, pri, csv_file) in enumerate(ISSUES):
            ws.write_formula(r, i + 2,
                '=COUNTIFS(A%d:A1048576,"%s",%s%d:%s1048576,A%d)' % (
                    T3_XL, label, PT1_COL, T3_XL, PT1_COL, r + 1),
                fmts["num"])

    # TABLE 3
    ws.write(T3_LBL, 0, "Table 3", fmts["lbl"])
    ws.set_row(T3_HDR, 30)
    for ci, h in enumerate(T3_HDRS):
        ws.write(T3_HDR, ci, h, fmts["t3_hdr_l"] if ci in LFT else fmts["t3_hdr"])

    for ri, row in enumerate(all_rows):
        r = T3_DATA + ri
        ws.set_row(r, 14.5)
        for ci, h in enumerate(T3_HDRS):
            v = row.get(h)
            if h in ("Impressions", "Clicks", "Organic Sessions",
                     "Errors", "Warnings", "Total Types", "Unique Types"):
                ws.write(r, ci, safe_num(v), fmts["t3_cell"])
            else:
                ws.write(r, ci, _clean(v) if v else "-",
                         fmts["t3_cell_l"] if ci in LFT else fmts["t3_cell"])

    wb.close()
    xlsx_bytes = buf.getvalue()

    # Pass 2: openpyxl rich-text summary
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
        tmp.write(xlsx_bytes)

    try:
        _apply_openpyxl(tmp_path, T3_HDR)
        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)


def _apply_openpyxl(path, t3_hdr_0idx):
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText

    wb  = load_workbook(path)
    ws  = wb.worksheets[0]
    bld = InlineFont(b=True,  rFont="Calibri", sz=10, color="FF000000")
    nrm = InlineFont(b=False, rFont="Calibri", sz=10, color="FF000000")
    bdr = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    gray = PatternFill(fill_type="solid", fgColor="FFD9D9D9")

    # Rich-text summary A1:H11
    try:
        ws.merge_cells("A1:H11")
    except Exception:
        pass
    c = ws["A1"]
    blocks = [TextBlock(bld if bold else nrm, txt) for bold, txt in SUMMARY_PAIRS]
    c.value     = CellRichText(*blocks)
    c.font      = Font(name="Calibri", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border    = bdr

    # Gray fill on T3 header row
    t3_xl = t3_hdr_0idx + 1
    max_col = 0
    for cell in ws[t3_xl]:
        if cell.value is not None:
            max_col = cell.column
    if max_col > 0:
        for col in range(1, max_col + 1):
            cell = ws.cell(row=t3_xl, column=col)
            cell.fill   = gray
            cell.border = bdr
            cell.font   = Font(name="Calibri", size=9, bold=True)

    wb.save(path)
