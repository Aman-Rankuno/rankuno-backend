import os
import io
import zipfile
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from app.config import settings
from app.services.rulebook import load_rulebook, classify_url

BLACK_FONT = Font(name="Arial", size=9, color="FF000000")
PRIORITY_ORDER = {"High": 0, "Medium": 1, "Low": 2}

T1_DATA_ROW = 12
T1_PCT_ROW = 13
T2_DATA_START_ROW = 19
T2_DATA_END_ROW = 27
T3_DATA_START_ROW = 30

ISSUE_KEYS = [
    "Uppercase", "Underscores", "Parameters",
    "Multiple Slashes", "Repetitive Path", "Contains Space"
]

ISSUE_PRIORITY = {
    "Uppercase": "Low", "Underscores": "Low", "Parameters": "Medium",
    "Multiple Slashes": "Medium", "Repetitive Path": "High", "Contains Space": "Low",
}

ISSUE_CSV = {
    "Uppercase": "url_uppercase.csv",
    "Underscores": "url_underscores.csv",
    "Parameters": "url_parameters.csv",
    "Multiple Slashes": "url_multiple_slashes.csv",
    "Repetitive Path": "url_repetitive_path.csv",
    "Contains Space": "url_contains_space.csv",
}

T1_ISSUE_START_COL = 2
T2_THEME_COL = 1
T2_PRIORITY_COL = 2
T2_ISSUE_START_COL = 3

T3_COLS = [
    "Error Type", "Address", "Page Type", "Page Theme 1", "Page Theme 2",
    "Content Type", "Status Code", "Indexability",
    "Impressions", "Clicks", "Organic Sessions"
]

KEEP_FROM_TEMPLATE = {'xl/drawings/drawing1.xml'}
DRAWING_CT = '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
DRAWING_REL = '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'
EMPTY_RELS = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>'


def _get_page_type(content_type: str) -> str:
    ct = str(content_type).lower()
    if "html" in ct:
        return "HTML"
    if "pdf" in ct:
        return "PDF"
    if ct.startswith("image/"):
        return "Image"
    if "css" in ct:
        return "CSS"
    if "javascript" in ct:
        return "JavaScript"
    if ct and ct != "nan":
        return "Other"
    return ""


def _merge_with_drawings(template_path: str, openpyxl_buf: io.BytesIO) -> bytes:
    openpyxl_buf.seek(0)

    # Step 1: combine openpyxl output + template drawing files
    step1_buf = io.BytesIO()
    with zipfile.ZipFile(template_path, 'r') as t_zip:
        with zipfile.ZipFile(openpyxl_buf, 'r') as o_zip:
            with zipfile.ZipFile(step1_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:
                for item in o_zip.namelist():
                    out_zip.writestr(item, o_zip.read(item))
                for item in KEEP_FROM_TEMPLATE:
                    if item in t_zip.namelist():
                        out_zip.writestr(item, t_zip.read(item))
                for item in t_zip.namelist():
                    if 'drawings/_rels' in item and item not in o_zip.namelist():
                        out_zip.writestr(item, t_zip.read(item))

    step1_buf.seek(0)
    step1_bytes = step1_buf.read()

    # Step 2: read content types and sheet rels, patch to include drawing refs
    with zipfile.ZipFile(io.BytesIO(step1_bytes), 'r') as z:
        ct = z.read('[Content_Types].xml').decode('utf-8')
        try:
            sheet_rels = z.read('xl/worksheets/_rels/sheet1.xml.rels').decode('utf-8')
        except KeyError:
            sheet_rels = EMPTY_RELS

    if 'drawing1' not in ct:
        ct = ct.replace('</Types>', DRAWING_CT + '</Types>')
    if 'drawing1' not in sheet_rels:
        sheet_rels = sheet_rels.replace('</Relationships>', DRAWING_REL + '</Relationships>')

    # Step 3: write final zip with patched files
    final_buf = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(step1_bytes), 'r') as z:
        with zipfile.ZipFile(final_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:
            for item in z.namelist():
                if item == '[Content_Types].xml':
                    out_zip.writestr(item, ct.encode('utf-8'))
                elif item == 'xl/worksheets/_rels/sheet1.xml.rels':
                    out_zip.writestr(item, sheet_rels.encode('utf-8'))
                else:
                    out_zip.writestr(item, z.read(item))
            # Write sheet rels if not already present
            if 'xl/worksheets/_rels/sheet1.xml.rels' not in z.namelist():
                out_zip.writestr('xl/worksheets/_rels/sheet1.xml.rels', sheet_rels.encode('utf-8'))

    final_buf.seek(0)
    return final_buf.read()


def build_url_issues_masterfile(crawl_id: str, domain: str, report_path: str) -> bytes:
    template_path = os.path.join(settings.TEMPLATES_DIR, "URL Issues.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    rulebook = load_rulebook(domain)

    def load_csv(filename):
        path = os.path.join(report_path, filename)
        if os.path.exists(path):
            try:
                return pd.read_csv(path, encoding="utf-8", low_memory=False)
            except Exception:
                return pd.DataFrame()
        return pd.DataFrame()

    df_internal_all = load_csv("internal_all.csv")
    df_gsc = load_csv("search_console_all.csv")
    df_ga = load_csv("analytics_all.csv")

    total_crawled = len(df_internal_all) if not df_internal_all.empty else 0

    gsc_map = {}
    if not df_gsc.empty:
        a = next((c for c in df_gsc.columns if c.lower() == "address"), None)
        imp = next((c for c in df_gsc.columns if "impression" in c.lower()), None)
        clk = next((c for c in df_gsc.columns if "click" in c.lower()), None)
        if a:
            for _, r in df_gsc.iterrows():
                gsc_map[str(r[a])] = {
                    "impressions": r.get(imp, 0) if imp else 0,
                    "clicks": r.get(clk, 0) if clk else 0,
                }

    ga_map = {}
    if not df_ga.empty:
        a = next((c for c in df_ga.columns if c.lower() == "address"), None)
        s = next((c for c in df_ga.columns if "session" in c.lower()), None)
        if a and s:
            for _, r in df_ga.iterrows():
                ga_map[str(r[a])] = r.get(s, 0)

    # Build internal_all lookup
    internal_map = {}
    if not df_internal_all.empty:
        def gc(df, *names):
            for n in names:
                m = next((c for c in df.columns if c.lower() == n.lower()), None)
                if m:
                    return m
            return None
        a_col = gc(df_internal_all, "Address")
        ct_col = gc(df_internal_all, "Content Type")
        sc_col = gc(df_internal_all, "Status Code")
        idx_col = gc(df_internal_all, "Indexability")
        if a_col:
            for _, r in df_internal_all.iterrows():
                internal_map[str(r[a_col])] = {
                    "content_type": str(r.get(ct_col, "")) if ct_col else "",
                    "status_code": str(r.get(sc_col, "")) if sc_col else "",
                    "indexability": str(r.get(idx_col, "")) if idx_col else "",
                }

    issue_counts = {k: 0 for k in ISSUE_KEYS}
    all_rows = []

    for issue_key, csv_file in ISSUE_CSV.items():
        df = load_csv(csv_file)
        if df.empty:
            continue

        def gc2(df, *names):
            for n in names:
                m = next((c for c in df.columns if c.lower() == n.lower()), None)
                if m:
                    return m
            return None

        addr_col = gc2(df, "Address")
        ct_col2 = gc2(df, "Content Type")
        sc_col2 = gc2(df, "Status Code")
        idx_col2 = gc2(df, "Indexability")

        for _, row in df.iterrows():
            url = str(row.get(addr_col, "")) if addr_col else ""
            int_data = internal_map.get(url, {})
            content_type = str(row.get(ct_col2, "")) if ct_col2 else int_data.get("content_type", "")
            status_code = str(row.get(sc_col2, "")) if sc_col2 else int_data.get("status_code", "")
            indexability = str(row.get(idx_col2, "")) if idx_col2 else int_data.get("indexability", "")
            page_type = _get_page_type(content_type)

            # Filter: only Indexable, 200, HTML
            if status_code != "200":
                continue
            if indexability.lower() != "indexable":
                continue
            if page_type != "HTML":
                continue

            issue_counts[issue_key] += 1
            page_theme1, page_theme2, _, priority = classify_url(url, rulebook)
            gsc = gsc_map.get(url, {})

            all_rows.append({
                "Error Type": issue_key,
                "Address": url,
                "Page Type": page_type,
                "Page Theme 1": page_theme1,
                "Page Theme 2": page_theme2,
                "Content Type": content_type,
                "Status Code": status_code,
                "Indexability": indexability,
                "Impressions": gsc.get("impressions", 0),
                "Clicks": gsc.get("clicks", 0),
                "Organic Sessions": ga_map.get(url, 0),
                "_priority": priority,
            })

    df_table3 = pd.DataFrame(all_rows) if all_rows else pd.DataFrame(columns=T3_COLS + ["_priority"])
    if not df_table3.empty:
        df_table3 = df_table3.sort_values("Impressions", ascending=False).reset_index(drop=True)

    theme_counts = {}
    for _, row in df_table3.iterrows():
        theme = row["Page Theme 1"] or "Others"
        priority = row.get("_priority", "Low")
        if theme not in theme_counts:
            theme_counts[theme] = {"total": 0, "_priority": priority}
            for k in ISSUE_KEYS:
                theme_counts[theme][k] = 0
        else:
            if PRIORITY_ORDER.get(priority, 2) < PRIORITY_ORDER.get(theme_counts[theme]["_priority"], 2):
                theme_counts[theme]["_priority"] = priority
        theme_counts[theme]["total"] += 1
        issue = str(row["Error Type"])
        if issue in theme_counts[theme]:
            theme_counts[theme][issue] += 1

    sorted_themes = sorted(
        theme_counts.items(),
        key=lambda x: PRIORITY_ORDER.get(x[1].get("_priority", "Low"), 2)
    )

    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb.active

    # Reset font color for data rows
    for r in list(range(T1_DATA_ROW, T1_PCT_ROW + 1)) + list(range(T2_DATA_START_ROW, T2_DATA_END_ROW + 1)) + list(range(T3_DATA_START_ROW, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).font = BLACK_FONT

    # Fill Table 1
    for i, issue in enumerate(ISSUE_KEYS):
        col = T1_ISSUE_START_COL + i
        count = issue_counts[issue]
        ws.cell(T1_DATA_ROW, col).value = count if count > 0 else None
        ws.cell(T1_DATA_ROW, col).font = BLACK_FONT
        pct = f"{round(count / total_crawled * 100, 1)}%" if total_crawled > 0 and count > 0 else None
        ws.cell(T1_PCT_ROW, col).value = pct
        ws.cell(T1_PCT_ROW, col).font = BLACK_FONT

    # Clear and fill Table 2
    for r in range(T2_DATA_START_ROW, T2_DATA_END_ROW + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    for row_offset, (theme, counts) in enumerate(sorted_themes):
        r = T2_DATA_START_ROW + row_offset
        total_theme = counts["total"]
        ws.cell(r, T2_THEME_COL).value = theme
        ws.cell(r, T2_THEME_COL).font = BLACK_FONT
        ws.cell(r, T2_PRIORITY_COL).value = counts.get("_priority", "Low")
        ws.cell(r, T2_PRIORITY_COL).font = BLACK_FONT
        for i, issue in enumerate(ISSUE_KEYS):
            cnt = counts.get(issue, 0)
            val = f"{cnt} ({round(cnt / total_theme * 100)}%)" if cnt > 0 and total_theme > 0 else None
            cell = ws.cell(r, T2_ISSUE_START_COL + i)
            cell.value = val
            cell.font = BLACK_FONT

    # Clear and fill Table 3
    for r in range(T3_DATA_START_ROW, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None

    for row_offset, (_, row) in enumerate(df_table3.iterrows()):
        r = T3_DATA_START_ROW + row_offset
        for col_offset, col_name in enumerate(T3_COLS):
            val = row.get(col_name, "")
            if pd.isna(val) or val == "":
                val = None
            cell = ws.cell(r, 1 + col_offset)
            cell.value = val
            cell.font = BLACK_FONT

    openpyxl_buf = io.BytesIO()
    wb.save(openpyxl_buf)
    return _merge_with_drawings(template_path, openpyxl_buf)